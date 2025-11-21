#!/usr/bin/env python3
"""
===============================================================================
VTOL PERFORMANCE ANALYZER v4.0 - PROFESSIONAL GUI
===============================================================================

Full-featured Tkinter GUI for professional drone performance analysis.

Features:
- Interactive parameter configuration
- Real-time analysis updates
- Custom plot generation (any parameter vs any parameter)
- Mission profile builder
- Multi-preset comparison
- Professional export tools (PDF, Excel, reports)

Usage:
    python vtol_analyzer_gui.py
    python vtol_performance_analyzer.py --gui

Version: 4.0.0
Date: 2025-01-20
===============================================================================
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import sys
import os
from pathlib import Path

# Matplotlib for embedded plots
import matplotlib
matplotlib.use('TkAgg')
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
import matplotlib.pyplot as plt

# Import core functionality
try:
    from analyzer import (
        AircraftConfiguration,
        PerformanceCalculator,
        ReportGenerator
    )
    from presets import PresetManager, get_preset
except ImportError as e:
    print(f"Error: Could not import required modules: {e}")
    print("Make sure all modules are in the src/ directory")
    sys.exit(1)


# ===========================================================================
# MAIN APPLICATION CLASS
# ===========================================================================

class VTOLAnalyzerGUI(tk.Tk):
    """
    Main GUI application for VTOL Performance Analyzer v4.0

    Professional desktop application with:
    - 6 tabbed interfaces
    - Real-time parameter editing
    - Interactive plotting
    - Mission builder
    - Comparison tools
    - Export manager
    """

    def __init__(self):
        super().__init__()

        # Window setup
        self.title("VTOL Performance Analyzer v4.0 - Professional Edition")
        self.geometry("1400x900")
        self.minsize(1200, 700)

        # Application icon (if available)
        try:
            self.iconbitmap('vtol_icon.ico')
        except:
            pass  # Icon optional

        # Core application data
        self.preset_manager = PresetManager()
        self.current_preset_name = "baseline"
        self.current_config = None
        self.current_calc = None
        self.current_results = None

        # Settings
        self.auto_update = tk.BooleanVar(value=False)
        self.dark_mode = tk.BooleanVar(value=False)
        self.auto_save_enabled = tk.BooleanVar(value=True)

        # Session management
        self.session_file = ".vtol_analyzer_session.json"
        self.config_dir = os.path.join(os.path.expanduser("~"), ".vtol_analyzer")
        os.makedirs(self.config_dir, exist_ok=True)
        self.recent_configs = []
        self.max_recent = 10
        self.config_modified = False
        self.auto_save_timer_id = None

        # Parameter tooltips database
        self.param_tooltips = self.get_parameter_tooltips()

        # Parameter validation ranges
        self.param_ranges = self.get_parameter_ranges()

        # Validation status tracking
        self.param_validation_status = {}
        self.overall_validation_valid = False

        # Initialize UI
        self.create_styles()
        self.create_menu()
        self.create_main_interface()
        self.create_status_bar()

        # Load session state
        self.load_session_state()

        # Start auto-save timer
        if self.auto_save_enabled.get():
            self.start_auto_save_timer()

        # Bind window close
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

    def get_parameter_tooltips(self):
        """Comprehensive tooltip database for all parameters"""
        return {
            # Basic Parameters
            "total_takeoff_weight_kg": "Total aircraft weight including battery, payload, and airframe.\nTypical range: 2-10 kg for small VTOL",
            "wingspan_m": "Total wing span from tip to tip.\nAffects lift generation and L/D ratio in forward flight",
            "wing_chord_m": "Average wing chord (width from leading to trailing edge).\nDetermines wing area with span",
            "field_elevation_m": "Operating altitude above sea level (affects air density).\nHigher altitude = less dense air = more power needed",

            # Tailsitter-Specific
            "control_power_base_w": "Base power for control surfaces and stabilization.\nHigher for aggressive flight or turbulent conditions",
            "control_power_speed_factor": "Additional control power per m/s airspeed.\nAccounts for increased control demands at higher speeds",
            "cd0_motor_nacelles": "Parasite drag from motors and nacelles.\n0.020-0.035 typical for clean installations",
            "cd0_fuselage_base": "Fuselage base drag coefficient.\n0.005-0.015 depending on streamlining",
            "cd0_landing_gear": "Drag from landing gear (if any).\n0.008-0.015 for fixed gear, 0.003 for retractable",
            "cd0_interference": "Drag from component interference (wing-body, etc).\n0.010-0.020 typical for integrated designs",

            # Transition Parameters
            "transition_duration_s": "Time to transition between hover and forward flight.\nLonger = smoother but more energy, Shorter = quicker but rougher",
            "avg_transition_power_w": "Average electrical power during transition.\nTypically 1.3-1.8x forward flight power",
            "transition_forward_duration_s": "Duration of hover-to-forward flight transition.\n10-20 seconds typical for smooth transition",
            "transition_forward_power_factor": "Power multiplier during forward transition.\n1.5-2.5× cruise power typical",
            "transition_back_duration_s": "Duration of forward-to-hover transition.\n8-15 seconds, usually faster than forward transition",
            "transition_back_power_factor": "Power multiplier during back transition.\n1.2-2.0× cruise power typical",

            # Propulsion
            "battery_capacity_mah": "Total battery capacity in milliamp-hours.\nHigher = longer flight time but more weight",
            "battery_voltage_v": "Nominal battery voltage.\nCommon: 14.8V (4S), 22.2V (6S), 29.6V (8S)",
            "battery_weight_fraction": "Battery weight as fraction of total weight.\n0.25-0.35 typical for good endurance",
            "battery_min_voltage_fraction": "Minimum safe voltage as fraction of nominal.\n0.80 = land at 80% voltage for battery longevity",
            "motor_kv": "Motor velocity constant (RPM per volt).\nLower KV = more torque, better for larger props",
            "propeller_diameter_in": "Propeller diameter in inches.\nLarger = more efficient but slower response",
            "propeller_pitch_in": "Propeller pitch in inches.\nHigher pitch = more thrust at high speed",
            "hover_efficiency": "Propeller efficiency in hover (vertical).\n0.50-0.70 typical for multirotor props",
            "cruise_efficiency": "Propeller efficiency in forward flight.\n0.75-0.85 typical for fixed-pitch cruise props",
            "motor_efficiency": "Motor electrical-to-mechanical efficiency.\n0.80-0.90 for quality brushless motors",
            "esc_efficiency": "Electronic Speed Controller efficiency.\n0.92-0.97 for modern ESCs",
            "prop_efficiency_lowspeed": "Propeller efficiency at low speed (hover/climb).\n60-75% typical for multirotor-style props",
            "prop_efficiency_highspeed": "Propeller efficiency at high speed (cruise).\n50-65% for fixed-pitch compromises, 75-85% for optimized cruise props",
            "motor_efficiency_peak": "Peak motor efficiency.\n75-90% depending on motor quality and operating point",

            # Auxiliary Systems
            "avionics_power_w": "Power consumed by flight controller, GPS, telemetry.\n4-8W typical for autopilot systems",
            "payload_power_w": "Power for cameras, sensors, or other payload.\n0-20W depending on equipment",
            "heater_power_w": "Battery heater power for cold weather operations.\n0-15W, only when needed",

            # Geometry Parameters (v4.1)
            "fuselage_length_m": "Total fuselage length from nose to tail.\nAffects parasite drag and structural weight.\n1.0-1.5m typical for small tailsitters",
            "fuselage_diameter_m": "Fuselage body diameter (cylindrical approximation).\nAffects frontal drag and internal volume.\n0.08-0.12m typical for small UAVs",
            "num_tail_fins": "Number of vertical stabilizers for yaw/pitch control.\n3 fins (120° apart) or 4 fins (90° apart) typical for tailsitters",
            "tail_fin_chord_m": "Root chord length of tail fin.\nAffects control authority and drag.\n0.04-0.06m typical for small aircraft",
            "tail_fin_span_m": "Vertical height of tail fin from root to tip.\nAffects stability and control power.\n0.12-0.20m typical",
            "tail_fin_position_m": "Distance from center of gravity to fin root (aft direction).\nFarther aft = better control leverage but more structural weight.\n0.4-0.6m typical",
            "tail_fin_thickness_ratio": "Airfoil thickness-to-chord ratio (symmetric airfoil).\n0.12 = NACA 0012, 0.10 = NACA 0010.\nThicker = stronger but more drag",
            "tail_fin_taper_ratio": "Ratio of tip chord to root chord.\n1.0 = rectangular, 0.7 = moderate taper, 0.5 = strong taper.\nTaper reduces weight and induced drag",
            "motor_spacing_m": "Distance between motors in quad configuration.\nAffects control authority and structural loads.\n0.4-0.6m typical for small quads",
            "num_motors": "Number of motors for VTOL thrust.\n4 = quadcopter, 6 = hexacopter.\nCurrently only 4 motors supported",
        }

    def get_parameter_ranges(self):
        """Parameter validation ranges (min, max, warning_min, warning_max)"""
        return {
            # Basic Parameters
            "total_takeoff_weight_kg": (1.0, 20.0, 2.0, 12.0),
            "wingspan_m": (0.5, 5.0, 1.0, 3.0),
            "wing_chord_m": (0.05, 0.5, 0.1, 0.3),
            "field_elevation_m": (0.0, 4000.0, 0.0, 2500.0),

            # Tailsitter-Specific
            "control_power_base_w": (20.0, 150.0, 30.0, 100.0),
            "control_power_speed_factor": (1.0, 15.0, 2.0, 10.0),
            "cd0_motor_nacelles": (0.010, 0.060, 0.020, 0.045),
            "cd0_fuselage_base": (0.003, 0.025, 0.005, 0.015),
            "cd0_landing_gear": (0.003, 0.030, 0.008, 0.020),
            "cd0_interference": (0.005, 0.035, 0.010, 0.025),

            # Transition Parameters
            "transition_forward_duration_s": (5.0, 30.0, 10.0, 20.0),
            "transition_forward_power_factor": (1.2, 3.0, 1.5, 2.5),
            "transition_back_duration_s": (4.0, 25.0, 8.0, 15.0),
            "transition_back_power_factor": (1.0, 2.5, 1.2, 2.0),

            # Propulsion
            "battery_capacity_mah": (2000.0, 30000.0, 5000.0, 20000.0),
            "battery_voltage_v": (7.4, 50.0, 11.1, 29.6),
            "battery_weight_fraction": (0.15, 0.50, 0.25, 0.40),
            "battery_min_voltage_fraction": (0.70, 0.95, 0.75, 0.85),
            "motor_kv": (100.0, 2000.0, 300.0, 1200.0),
            "propeller_diameter_in": (5.0, 20.0, 8.0, 16.0),
            "propeller_pitch_in": (3.0, 15.0, 5.0, 12.0),
            "hover_efficiency": (0.40, 0.80, 0.50, 0.70),
            "cruise_efficiency": (0.60, 0.90, 0.75, 0.85),
            "motor_efficiency": (0.70, 0.95, 0.80, 0.90),
            "esc_efficiency": (0.85, 0.98, 0.92, 0.97),
            "prop_efficiency_lowspeed": (0.40, 0.85, 0.60, 0.75),
            "prop_efficiency_highspeed": (0.40, 0.90, 0.50, 0.85),
            "motor_efficiency_peak": (0.70, 0.95, 0.75, 0.90),

            # Auxiliary Systems
            "avionics_power_w": (2.0, 20.0, 4.0, 10.0),
            "payload_power_w": (0.0, 50.0, 0.0, 20.0),
            "heater_power_w": (0.0, 30.0, 0.0, 15.0),

            # Geometry Parameters (v4.1)
            # Format: (hard_min, hard_max, warning_min, warning_max)
            "fuselage_length_m": (0.5, 2.5, 0.8, 2.0),
            "fuselage_diameter_m": (0.03, 0.25, 0.05, 0.20),
            "num_tail_fins": (3.0, 4.0, 3.0, 4.0),  # Only 3 or 4 supported
            "tail_fin_chord_m": (0.02, 0.15, 0.03, 0.10),
            "tail_fin_span_m": (0.05, 0.40, 0.10, 0.30),
            "tail_fin_position_m": (0.2, 1.0, 0.3, 0.8),
            "tail_fin_thickness_ratio": (0.06, 0.18, 0.08, 0.15),
            "tail_fin_taper_ratio": (0.4, 1.0, 0.5, 1.0),
            "motor_spacing_m": (0.2, 1.0, 0.3, 0.8),
            "num_motors": (4.0, 4.0, 4.0, 4.0),  # Only 4 motors supported currently
        }

    def validate_parameter(self, param_name, value_str):
        """Validate a single parameter and return status"""
        if param_name not in self.param_ranges:
            return 'unknown', None

        try:
            value = float(value_str)
            min_val, max_val, warn_min, warn_max = self.param_ranges[param_name]

            # Check hard limits
            if value < min_val or value > max_val:
                return 'invalid', f"Out of range: {min_val}-{max_val}"

            # Check warning range
            if value < warn_min or value > warn_max:
                return 'warning', f"Unusual value (typical: {warn_min}-{warn_max})"

            # Valid
            return 'valid', None

        except ValueError:
            return 'invalid', "Must be a number"

    def bind_validation(self, param_name, entry_widget):
        """Bind real-time validation to an entry widget"""
        def on_change(event=None):
            value_str = entry_widget.get()

            # Skip if empty (during loading)
            if not value_str:
                entry_widget.config(background='white')
                self.param_validation_status[param_name] = 'empty'
                self.update_overall_validation()
                return

            # Validate
            status, message = self.validate_parameter(param_name, value_str)
            self.param_validation_status[param_name] = status

            # Visual feedback
            colors = {
                'valid': '#D4EDDA',      # Light green
                'warning': '#FFF3CD',    # Light yellow
                'invalid': '#F8D7DA',    # Light red
                'empty': 'white',
                'unknown': 'white',
            }
            entry_widget.config(background=colors.get(status, 'white'))

            # Update overall validation
            self.update_overall_validation()

        # Bind to key release
        entry_widget.bind('<KeyRelease>', on_change)
        entry_widget.bind('<FocusOut>', on_change)

        # Initial validation
        on_change()

    def update_overall_validation(self):
        """Update overall validation status and enable/disable Run Analysis"""
        # Check if any parameter is invalid
        has_invalid = any(status == 'invalid' for status in self.param_validation_status.values())
        has_empty = any(status == 'empty' for status in self.param_validation_status.values())

        # Update overall status
        if has_invalid or has_empty:
            self.overall_validation_valid = False
        else:
            self.overall_validation_valid = True

        # Update Run Analysis button state (if it exists)
        if hasattr(self, 'run_analysis_button'):
            if self.overall_validation_valid:
                self.run_analysis_button.config(state='normal')
            else:
                self.run_analysis_button.config(state='disabled')

    # -----------------------------------------------------------------------
    # STYLING
    # -----------------------------------------------------------------------

    def create_styles(self):
        """Create professional ttk styles"""
        style = ttk.Style()

        # Try to use a modern theme
        try:
            style.theme_use('clam')  # Modern, flat theme
        except:
            pass

        # Custom colors
        self.colors = {
            'primary': '#2C3E50',
            'secondary': '#3498DB',
            'success': '#27AE60',
            'warning': '#F39C12',
            'danger': '#E74C3C',
            'background': '#ECF0F1',
            'text': '#2C3E50',
            'accent': '#9B59B6',
        }

        # Configure styles
        style.configure('Title.TLabel', font=('Arial', 14, 'bold'), foreground=self.colors['primary'])
        style.configure('Heading.TLabel', font=('Arial', 11, 'bold'))
        style.configure('Status.TLabel', font=('Arial', 9))
        style.configure('Primary.TButton', foreground='white', background=self.colors['primary'])

        # Tab style
        style.configure('TNotebook', background=self.colors['background'])
        style.configure('TNotebook.Tab', padding=[20, 10], font=('Arial', 10))

    # -----------------------------------------------------------------------
    # MENU BAR
    # -----------------------------------------------------------------------

    def create_menu(self):
        """Create application menu bar"""
        menubar = tk.Menu(self)
        self.config(menu=menubar)

        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="New Analysis", command=self.new_analysis, accelerator="Ctrl+N")
        file_menu.add_command(label="Open Configuration...", command=self.open_config, accelerator="Ctrl+O")
        file_menu.add_command(label="Save Configuration...", command=self.save_config, accelerator="Ctrl+S")
        file_menu.add_separator()
        file_menu.add_command(label="Export Report...", command=self.export_report)
        file_menu.add_command(label="Export All Data...", command=self.export_all_data)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.on_closing, accelerator="Ctrl+Q")

        # View menu
        view_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="View", menu=view_menu)
        view_menu.add_checkbutton(label="Auto Update Analysis", variable=self.auto_update, command=self.toggle_auto_update)
        view_menu.add_checkbutton(label="Auto-Save Configuration", variable=self.auto_save_enabled, command=self.toggle_auto_save)
        view_menu.add_checkbutton(label="Dark Mode", variable=self.dark_mode, command=self.toggle_dark_mode)
        view_menu.add_separator()
        view_menu.add_command(label="Refresh Results", command=self.refresh_results, accelerator="F5")

        # Tools menu
        tools_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Tools", menu=tools_menu)
        tools_menu.add_command(label="Mission Builder", command=lambda: self.notebook.select(3))
        tools_menu.add_command(label="Comparison Tool", command=lambda: self.notebook.select(4))
        tools_menu.add_command(label="Parameter Optimizer", command=self.optimize_parameters)
        tools_menu.add_separator()
        tools_menu.add_command(label="Run Script Mode...", command=self.run_script_mode)

        # Help menu
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="Quick Start Guide", command=self.show_quick_start)
        help_menu.add_command(label="Documentation", command=self.show_documentation)
        help_menu.add_command(label="Parameter Guide", command=self.show_parameter_guide)
        help_menu.add_separator()
        help_menu.add_command(label="About", command=self.show_about)

        # Keyboard shortcuts
        self.bind_all("<Control-n>", lambda e: self.new_analysis())
        self.bind_all("<Control-o>", lambda e: self.open_config())
        self.bind_all("<Control-s>", lambda e: self.save_config())
        self.bind_all("<Control-q>", lambda e: self.on_closing())
        self.bind_all("<F5>", lambda e: self.refresh_results())

    # -----------------------------------------------------------------------
    # MAIN INTERFACE
    # -----------------------------------------------------------------------

    def create_main_interface(self):
        """Create main tabbed notebook interface"""
        # Create notebook (tabbed interface)
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill='both', expand=True, padx=5, pady=5)

        # Tab 1: Configuration
        self.tab_config = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_config, text="  Configuration  ")
        self.create_config_tab()

        # Tab 2: Analysis Results
        self.tab_results = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_results, text="  Analysis Results  ")
        self.create_results_tab()

        # Tab 3: Interactive Plots
        self.tab_plots = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_plots, text="  Interactive Plots  ")
        self.create_plots_tab()

        # Tab 4: Mission Builder
        self.tab_mission = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_mission, text="  Mission Builder  ")
        self.create_mission_tab()

        # Tab 5: Comparison
        self.tab_comparison = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_comparison, text="  Comparison  ")
        self.create_comparison_tab()

        # Tab 6: Export Manager
        self.tab_export = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_export, text="  Export Manager  ")
        self.create_export_tab()

        # Tab 7: Design Schematic (v4.1)
        self.tab_schematic = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_schematic, text="  Design Schematic  ")
        self.create_schematic_tab()

    # -----------------------------------------------------------------------
    # TAB 1: CONFIGURATION
    # -----------------------------------------------------------------------

    def create_config_tab(self):
        """Create configuration tab with preset selector and parameter editor"""
        # Top section: Preset selector
        top_frame = ttk.Frame(self.tab_config)
        top_frame.pack(fill='x', padx=10, pady=10)

        ttk.Label(top_frame, text="Preset:", font=('Arial', 11, 'bold')).pack(side='left', padx=5)

        self.preset_var = tk.StringVar(value="baseline")
        preset_combo = ttk.Combobox(top_frame, textvariable=self.preset_var, width=50, state='readonly')
        preset_combo['values'] = [
            f"{name}: {self.preset_manager.get_preset_description(name)}"
            for name in self.preset_manager.list_presets()
        ]
        preset_combo.current(1)  # baseline
        preset_combo.pack(side='left', padx=5)
        preset_combo.bind('<<ComboboxSelected>>', self.on_preset_selected)

        ttk.Button(top_frame, text="Load", command=self.load_selected_preset).pack(side='left', padx=2)
        ttk.Button(top_frame, text="Save As...", command=self.save_preset_as).pack(side='left', padx=2)
        ttk.Button(top_frame, text="Reset", command=self.reset_config).pack(side='left', padx=2)
        ttk.Button(top_frame, text="Apply Changes", command=self.apply_config, style='Primary.TButton').pack(side='left', padx=10)

        # Main scrollable frame for parameters
        canvas = tk.Canvas(self.tab_config, bg='white')
        scrollbar = ttk.Scrollbar(self.tab_config, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
        scrollbar.pack(side="right", fill="y", pady=10, padx=(0, 10))

        # Store parameter widgets
        self.param_widgets = {}

        # Create parameter sections
        self.create_basic_params_section(scrollable_frame)
        self.create_tailsitter_params_section(scrollable_frame)
        self.create_transitions_params_section(scrollable_frame)
        self.create_propulsion_params_section(scrollable_frame)
        self.create_auxiliary_params_section(scrollable_frame)
        self.create_geometry_params_section(scrollable_frame)  # v4.1

        # Bottom action buttons
        bottom_frame = ttk.Frame(self.tab_config)
        bottom_frame.pack(fill='x', padx=10, pady=10)

        ttk.Button(bottom_frame, text="Validate Configuration", command=self.validate_config).pack(side='left', padx=5)

        # Store Run Analysis button for validation control
        self.run_analysis_button = ttk.Button(bottom_frame, text="Run Analysis", command=self.run_analysis, style='Primary.TButton')
        self.run_analysis_button.pack(side='left', padx=5)

        ttk.Button(bottom_frame, text="View Results →", command=lambda: self.notebook.select(1)).pack(side='left', padx=5)

    def create_basic_params_section(self, parent):
        """Create basic parameters section"""
        frame = ttk.LabelFrame(parent, text=" Basic Parameters ", padding=10)
        frame.pack(fill='x', padx=10, pady=5)

        params = [
            ("total_takeoff_weight_kg", "Total Weight", "kg", "1.0-20.0"),
            ("wingspan_m", "Wing Span", "m", "0.5-5.0"),
            ("wing_chord_m", "Wing Chord", "m", "0.05-0.5"),
            ("field_elevation_m", "Field Elevation", "m MSL", "0-4000"),
        ]

        for i, (param, label, unit, range_str) in enumerate(params):
            row = ttk.Frame(frame)
            row.pack(fill='x', pady=2)

            ttk.Label(row, text=f"{label}:", width=20).pack(side='left')
            entry = ttk.Entry(row, width=15)
            entry.pack(side='left', padx=5)
            ttk.Label(row, text=unit, width=10).pack(side='left')
            ttk.Label(row, text=f"({range_str})", font=('Arial', 8), foreground='gray').pack(side='left')

            self.param_widgets[param] = entry

            # Add tooltip if available
            if param in self.param_tooltips:
                self.create_tooltip(entry, self.param_tooltips[param])

            # Bind real-time validation
            self.bind_validation(param, entry)

    def create_tailsitter_params_section(self, parent):
        """Create tailsitter-specific parameters section"""
        frame = ttk.LabelFrame(parent, text=" Tailsitter-Specific (v3.0) ", padding=10)
        frame.pack(fill='x', padx=10, pady=5)

        params = [
            ("control_power_base_w", "Control Power Base", "W", "30-100", True),
            ("control_power_speed_factor", "Control Power Speed Factor", "W/(m/s)", "2-10", True),
            ("cd0_motor_nacelles", "CD0 Motor Nacelles", "-", "0.020-0.050", True),
            ("cd0_fuselage_base", "CD0 Fuselage Base", "-", "0.005-0.015", False),
            ("cd0_landing_gear", "CD0 Landing Gear", "-", "0.008-0.020", False),
            ("cd0_interference", "CD0 Interference", "-", "0.010-0.025", True),
        ]

        for param, label, unit, range_str, tunable in params:
            row = ttk.Frame(frame)
            row.pack(fill='x', pady=2)

            ttk.Label(row, text=f"{label}:", width=25).pack(side='left')
            entry = ttk.Entry(row, width=12)
            entry.pack(side='left', padx=5)
            ttk.Label(row, text=unit, width=12).pack(side='left')
            ttk.Label(row, text=f"({range_str})", font=('Arial', 8), foreground='gray').pack(side='left', padx=5)

            if tunable:
                ttk.Label(row, text="[TUNE]", foreground=self.colors['warning'], font=('Arial', 8, 'bold')).pack(side='left')

            self.param_widgets[param] = entry

            # Add tooltip if available
            if param in self.param_tooltips:
                self.create_tooltip(entry, self.param_tooltips[param])

            # Bind real-time validation
            self.bind_validation(param, entry)

    def create_transitions_params_section(self, parent):
        """Create transition parameters section (collapsible)"""
        # This will be a collapsible section - simplified for now
        frame = ttk.LabelFrame(parent, text=" Transitions ", padding=10)
        frame.pack(fill='x', padx=10, pady=5)

        params = [
            ("transition_forward_duration_s", "Forward Duration", "s", "10-20", True),
            ("transition_forward_power_factor", "Forward Power Factor", "×", "1.5-2.5", True),
            ("transition_back_duration_s", "Back Duration", "s", "8-15", False),
            ("transition_back_power_factor", "Back Power Factor", "×", "1.2-2.0", False),
        ]

        for param, label, unit, range_str, tunable in params:
            row = ttk.Frame(frame)
            row.pack(fill='x', pady=2)

            ttk.Label(row, text=f"{label}:", width=25).pack(side='left')
            entry = ttk.Entry(row, width=12)
            entry.pack(side='left', padx=5)
            ttk.Label(row, text=unit, width=12).pack(side='left')
            ttk.Label(row, text=f"({range_str})", font=('Arial', 8), foreground='gray').pack(side='left', padx=5)

            if tunable:
                ttk.Label(row, text="[MEASURE]", foreground=self.colors['secondary'], font=('Arial', 8, 'bold')).pack(side='left')

            self.param_widgets[param] = entry

            # Add tooltip if available
            if param in self.param_tooltips:
                self.create_tooltip(entry, self.param_tooltips[param])

            # Bind real-time validation
            self.bind_validation(param, entry)

    def create_propulsion_params_section(self, parent):
        """Create propulsion parameters section"""
        frame = ttk.LabelFrame(parent, text=" Propulsion Efficiency ", padding=10)
        frame.pack(fill='x', padx=10, pady=5)

        params = [
            ("prop_efficiency_lowspeed", "Prop Efficiency (Low Speed)", "%", "60-75"),
            ("prop_efficiency_highspeed", "Prop Efficiency (High Speed)", "%", "50-65"),
            ("motor_efficiency_peak", "Motor Efficiency", "%", "75-90"),
            ("esc_efficiency", "ESC Efficiency", "%", "88-95"),
        ]

        for param, label, unit, range_str in params:
            row = ttk.Frame(frame)
            row.pack(fill='x', pady=2)

            ttk.Label(row, text=f"{label}:", width=30).pack(side='left')
            entry = ttk.Entry(row, width=12)
            entry.pack(side='left', padx=5)
            ttk.Label(row, text=unit, width=8).pack(side='left')
            ttk.Label(row, text=f"({range_str})", font=('Arial', 8), foreground='gray').pack(side='left')

            self.param_widgets[param] = entry

            # Add tooltip if available
            if param in self.param_tooltips:
                self.create_tooltip(entry, self.param_tooltips[param])

            # Bind real-time validation
            self.bind_validation(param, entry)

    def create_auxiliary_params_section(self, parent):
        """Create auxiliary systems parameters section"""
        frame = ttk.LabelFrame(parent, text=" Auxiliary Systems ", padding=10)
        frame.pack(fill='x', padx=10, pady=5)

        params = [
            ("avionics_power_w", "Avionics Power", "W", "4-10"),
            ("payload_power_w", "Payload Power", "W", "0-20"),
            ("heater_power_w", "Heater Power", "W", "0-15"),
        ]

        for param, label, unit, range_str in params:
            row = ttk.Frame(frame)
            row.pack(fill='x', pady=2)

            ttk.Label(row, text=f"{label}:", width=20).pack(side='left')
            entry = ttk.Entry(row, width=12)
            entry.pack(side='left', padx=5)
            ttk.Label(row, text=unit, width=8).pack(side='left')
            ttk.Label(row, text=f"({range_str})", font=('Arial', 8), foreground='gray').pack(side='left')

            self.param_widgets[param] = entry

            # Add tooltip if available
            if param in self.param_tooltips:
                self.create_tooltip(entry, self.param_tooltips[param])

            # Bind real-time validation
            self.bind_validation(param, entry)

    def create_geometry_params_section(self, parent):
        """Create airframe geometry parameters section (v4.1)"""
        frame = ttk.LabelFrame(parent, text=" Airframe Geometry (v4.1) ", padding=10)
        frame.pack(fill='x', padx=10, pady=5)

        # Fuselage geometry
        ttk.Label(frame, text="Fuselage:", font=('Arial', 9, 'bold')).pack(anchor='w', pady=(5, 2))

        fuselage_params = [
            ("fuselage_length_m", "Fuselage Length", "m", "0.8-2.0"),
            ("fuselage_diameter_m", "Fuselage Diameter", "m", "0.05-0.20"),
        ]

        for param, label, unit, range_str in fuselage_params:
            row = ttk.Frame(frame)
            row.pack(fill='x', pady=2)

            ttk.Label(row, text=f"  {label}:", width=25).pack(side='left')
            entry = ttk.Entry(row, width=12)
            entry.pack(side='left', padx=5)
            ttk.Label(row, text=unit, width=8).pack(side='left')
            ttk.Label(row, text=f"({range_str})", font=('Arial', 8), foreground='gray').pack(side='left')

            self.param_widgets[param] = entry

            # Add tooltip if available
            if param in self.param_tooltips:
                self.create_tooltip(entry, self.param_tooltips[param])

            # Bind real-time validation
            self.bind_validation(param, entry)

        # Tail fin configuration
        ttk.Label(frame, text="Tail Fins:", font=('Arial', 9, 'bold')).pack(anchor='w', pady=(10, 2))

        tail_params = [
            ("num_tail_fins", "Number of Fins", "-", "3 or 4"),
            ("tail_fin_chord_m", "Fin Chord", "m", "0.03-0.10"),
            ("tail_fin_span_m", "Fin Span (Height)", "m", "0.10-0.30"),
            ("tail_fin_position_m", "Fin Position (aft of CG)", "m", "0.3-0.8"),
            ("tail_fin_thickness_ratio", "Airfoil Thickness Ratio", "-", "0.08-0.15"),
            ("tail_fin_taper_ratio", "Taper Ratio (tip/root)", "-", "0.5-1.0"),
        ]

        for param, label, unit, range_str in tail_params:
            row = ttk.Frame(frame)
            row.pack(fill='x', pady=2)

            ttk.Label(row, text=f"  {label}:", width=25).pack(side='left')
            entry = ttk.Entry(row, width=12)
            entry.pack(side='left', padx=5)
            ttk.Label(row, text=unit, width=8).pack(side='left')
            ttk.Label(row, text=f"({range_str})", font=('Arial', 8), foreground='gray').pack(side='left')

            self.param_widgets[param] = entry

            # Add tooltip if available
            if param in self.param_tooltips:
                self.create_tooltip(entry, self.param_tooltips[param])

            # Bind real-time validation
            self.bind_validation(param, entry)

        # Motor configuration
        ttk.Label(frame, text="Motors:", font=('Arial', 9, 'bold')).pack(anchor='w', pady=(10, 2))

        motor_params = [
            ("num_motors", "Number of Motors", "-", "4"),
            ("motor_spacing_m", "Motor Spacing", "m", "0.3-0.8"),
        ]

        for param, label, unit, range_str in motor_params:
            row = ttk.Frame(frame)
            row.pack(fill='x', pady=2)

            ttk.Label(row, text=f"  {label}:", width=25).pack(side='left')
            entry = ttk.Entry(row, width=12)
            entry.pack(side='left', padx=5)
            ttk.Label(row, text=unit, width=8).pack(side='left')
            ttk.Label(row, text=f"({range_str})", font=('Arial', 8), foreground='gray').pack(side='left')

            self.param_widgets[param] = entry

            # Add tooltip if available
            if param in self.param_tooltips:
                self.create_tooltip(entry, self.param_tooltips[param])

            # Bind real-time validation
            self.bind_validation(param, entry)

    # -----------------------------------------------------------------------
    # TAB 2: ANALYSIS RESULTS
    # -----------------------------------------------------------------------

    def create_results_tab(self):
        """Create analysis results display tab"""
        # Top toolbar
        toolbar = ttk.Frame(self.tab_results)
        toolbar.pack(fill='x', padx=10, pady=5)

        ttk.Label(toolbar, text="Performance Summary", style='Title.TLabel').pack(side='left')
        ttk.Button(toolbar, text="Export PDF", command=self.export_results_pdf).pack(side='right', padx=2)
        ttk.Button(toolbar, text="Copy Text", command=self.copy_results_text).pack(side='right', padx=2)
        ttk.Button(toolbar, text="Refresh", command=self.refresh_results).pack(side='right', padx=2)

        # Scrollable text area for results
        self.results_text = scrolledtext.ScrolledText(
            self.tab_results,
            wrap=tk.WORD,
            width=100,
            height=40,
            font=('Courier', 10)
        )
        self.results_text.pack(fill='both', expand=True, padx=10, pady=10)

        # Initial message
        self.results_text.insert('1.0', "No analysis run yet.\n\nGo to Configuration tab and click 'Run Analysis'.")
        self.results_text.config(state='disabled')

    # -----------------------------------------------------------------------
    # TAB 3: INTERACTIVE PLOTS
    # -----------------------------------------------------------------------

    def create_plots_tab(self):
        """Create interactive plotting tab with dynamic parameter selection"""
        # Initialize plot parameters list
        self.plot_params = []

        # Top section: Common Plots Gallery
        gallery_frame = ttk.LabelFrame(self.tab_plots, text=" 📊 Common Plots Gallery - Quick Click ", padding=10)
        gallery_frame.pack(fill='x', padx=10, pady=5)

        # Try to load common plots
        try:
            from plots import COMMON_PLOTS, PLOT_CATEGORIES
            self.common_plots_available = True

            # Create category sections
            for category, plot_ids in PLOT_CATEGORIES.items():
                cat_frame = ttk.LabelFrame(gallery_frame, text=f" {category} ", padding=5)
                cat_frame.pack(fill='x', pady=3)

                for plot_id in plot_ids:
                    plot_def = COMMON_PLOTS.get(plot_id)
                    if plot_def:
                        btn_text = f"{plot_def['icon']} {plot_def['name']}"
                        btn = ttk.Button(
                            cat_frame,
                            text=btn_text,
                            command=lambda pid=plot_id: self.load_common_plot(pid),
                            width=25
                        )
                        btn.pack(side='left', padx=3)

                        # Add tooltip with description
                        self.create_tooltip(btn, plot_def['description'])

        except ImportError:
            self.common_plots_available = False
            ttk.Label(gallery_frame, text="⚠ Common plots database not available",
                     foreground='orange').pack()

        # Custom plot configuration panel
        control_frame = ttk.LabelFrame(self.tab_plots, text=" ⚙ Custom Plot - Select Parameters ", padding=10)
        control_frame.pack(fill='x', padx=10, pady=10)

        # Instructions
        instructions = ttk.Label(
            control_frame,
            text="Add 2-3 parameters below. Plot shows: Last parameter vs First (if 2), or First vs Second,Third (if 3)",
            foreground='gray',
            font=('Arial', 9, 'italic')
        )
        instructions.pack(anchor='w', pady=(0, 10))

        # Dynamic parameter selector
        selector_frame = ttk.Frame(control_frame)
        selector_frame.pack(fill='x')

        # Container for parameter list
        self.params_container = ttk.Frame(selector_frame)
        self.params_container.pack(side='left', fill='both', expand=True)

        # Add parameter button
        add_btn_frame = ttk.Frame(selector_frame)
        add_btn_frame.pack(side='right', padx=10)

        ttk.Button(
            add_btn_frame,
            text="➕ Add Parameter",
            command=self.add_plot_parameter,
            style='Primary.TButton'
        ).pack()

        ttk.Label(add_btn_frame, text="(max 3)", font=('Arial', 8), foreground='gray').pack()

        # Action buttons
        button_frame = ttk.Frame(control_frame)
        button_frame.pack(fill='x', pady=10)

        ttk.Button(button_frame, text="📈 Generate Plot", command=self.generate_custom_plot,
                  style='Primary.TButton').pack(side='left', padx=5)
        ttk.Button(button_frame, text="Clear All", command=self.clear_all_plot_params).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Export PNG", command=self.export_plot_png).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Export CSV", command=self.export_plot_csv).pack(side='left', padx=5)

        # Matplotlib canvas
        self.plot_frame = ttk.Frame(self.tab_plots)
        self.plot_frame.pack(fill='both', expand=True, padx=10, pady=10)

        # Placeholder
        placeholder = ttk.Label(self.plot_frame,
                               text="👆 Click a common plot above or build your own below",
                               font=('Arial', 12))
        placeholder.pack(expand=True)

        # Pre-load 2 parameters to show interface
        self.after(100, lambda: self.add_plot_parameter("Speed (m/s)"))
        self.after(150, lambda: self.add_plot_parameter("Forward Flight Power (W)"))

    # -----------------------------------------------------------------------
    # TAB 4: MISSION BUILDER
    # -----------------------------------------------------------------------

    def create_mission_tab(self):
        """Create mission builder tab"""
        # Initialize mission segments list
        self.mission_segments = []

        # Top toolbar
        toolbar = ttk.Frame(self.tab_mission)
        toolbar.pack(fill='x', padx=10, pady=5)

        ttk.Label(toolbar, text="Mission Profile Builder", style='Title.TLabel').pack(side='left')

        # Add segment dropdown
        ttk.Label(toolbar, text="Add Segment:", font=('Arial', 10)).pack(side='right', padx=5)
        segment_types = ["Hover", "Cruise", "Transition Forward", "Transition Back"]
        self.segment_type_var = tk.StringVar(value="Hover")
        segment_combo = ttk.Combobox(toolbar, textvariable=self.segment_type_var, values=segment_types, width=20, state='readonly')
        segment_combo.pack(side='right', padx=5)
        ttk.Button(toolbar, text="Add", command=self.add_mission_segment, style='Primary.TButton').pack(side='right', padx=2)

        # Mission segments list
        segments_frame = ttk.LabelFrame(self.tab_mission, text=" Mission Segments ", padding=10)
        segments_frame.pack(fill='both', expand=True, padx=10, pady=5)

        # Scrollable canvas for segments
        canvas = tk.Canvas(segments_frame, height=300)
        scrollbar = ttk.Scrollbar(segments_frame, orient="vertical", command=canvas.yview)
        self.segments_frame_inner = ttk.Frame(canvas)

        self.segments_frame_inner.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=self.segments_frame_inner, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Mission summary panel
        summary_frame = ttk.LabelFrame(self.tab_mission, text=" Mission Summary ", padding=10)
        summary_frame.pack(fill='x', padx=10, pady=5)

        self.mission_summary_text = tk.StringVar(value="No segments added yet")
        summary_label = ttk.Label(summary_frame, textvariable=self.mission_summary_text, font=('Courier', 10))
        summary_label.pack()

        # Template selection frame
        template_frame = ttk.LabelFrame(self.tab_mission, text=" 📋 Mission Templates (Examples) ", padding=10)
        template_frame.pack(fill='x', padx=10, pady=5)

        # Try to import mission templates
        try:
            from mission_templates import get_template_names, DEFAULT_MISSION
            self.mission_templates_available = True

            ttk.Label(template_frame, text="Load Template:").pack(side='left', padx=5)

            template_names = get_template_names()
            self.template_var = tk.StringVar(value=template_names[0][1] if template_names else "")
            template_combo = ttk.Combobox(
                template_frame,
                textvariable=self.template_var,
                values=[name for _, name in template_names],
                width=40,
                state='readonly'
            )
            template_combo.pack(side='left', padx=5)

            ttk.Button(
                template_frame,
                text="Load Template",
                command=self.load_mission_template
            ).pack(side='left', padx=5)

            # Store template mapping
            self.template_map = {name: tid for tid, name in template_names}

        except ImportError:
            self.mission_templates_available = False
            ttk.Label(
                template_frame,
                text="Mission templates not available. Add mission_templates.py to enable.",
                foreground='gray',
                font=('Arial', 9, 'italic')
            ).pack(side='left', padx=5)

        # Action buttons
        action_frame = ttk.Frame(self.tab_mission)
        action_frame.pack(fill='x', padx=10, pady=10)

        ttk.Button(action_frame, text="Simulate Mission", command=self.simulate_mission, style='Primary.TButton').pack(side='left', padx=5)
        ttk.Button(action_frame, text="Clear All", command=self.clear_mission_segments).pack(side='left', padx=5)
        ttk.Button(action_frame, text="Save Mission", command=self.save_mission).pack(side='left', padx=5)
        ttk.Button(action_frame, text="Load Mission", command=self.load_mission).pack(side='left', padx=5)

        # Load default example mission on startup
        if self.mission_templates_available:
            self.after(100, lambda: self.load_mission_template(DEFAULT_MISSION))

    def add_mission_segment(self):
        """Add a mission segment to the list"""
        segment_type = self.segment_type_var.get()

        # Create segment frame
        segment_frame = ttk.Frame(self.segments_frame_inner, relief='ridge', borderwidth=2)
        segment_frame.pack(fill='x', pady=5, padx=5)

        # Segment number
        segment_num = len(self.mission_segments) + 1
        ttk.Label(segment_frame, text=f"{segment_num}.", font=('Arial', 10, 'bold')).grid(row=0, column=0, padx=5, pady=5)

        # Segment type label
        type_label = ttk.Label(segment_frame, text=segment_type, font=('Arial', 10))
        type_label.grid(row=0, column=1, sticky='w', padx=5)

        # Segment parameters
        segment_data = {'type': segment_type.lower().replace(' ', '_'), 'frame': segment_frame}

        if segment_type == "Hover":
            ttk.Label(segment_frame, text="Duration:").grid(row=0, column=2, padx=5)
            duration_var = tk.StringVar(value="60")
            ttk.Entry(segment_frame, textvariable=duration_var, width=10).grid(row=0, column=3, padx=5)
            ttk.Label(segment_frame, text="s").grid(row=0, column=4, sticky='w')
            segment_data['duration_s'] = duration_var

        elif segment_type == "Cruise":
            ttk.Label(segment_frame, text="Duration:").grid(row=0, column=2, padx=5)
            duration_var = tk.StringVar(value="600")
            ttk.Entry(segment_frame, textvariable=duration_var, width=10).grid(row=0, column=3, padx=5)
            ttk.Label(segment_frame, text="s").grid(row=0, column=4, sticky='w')

            ttk.Label(segment_frame, text="Speed:").grid(row=0, column=5, padx=5)
            speed_var = tk.StringVar(value="15.0")
            ttk.Entry(segment_frame, textvariable=speed_var, width=10).grid(row=0, column=6, padx=5)
            ttk.Label(segment_frame, text="m/s").grid(row=0, column=7, sticky='w')

            segment_data['duration_s'] = duration_var
            segment_data['speed_ms'] = speed_var

        elif "Transition" in segment_type:
            ttk.Label(segment_frame, text="(Auto-calculated)").grid(row=0, column=2, padx=5)

        # Control buttons
        btn_frame = ttk.Frame(segment_frame)
        btn_frame.grid(row=0, column=8, padx=5)

        ttk.Button(btn_frame, text="↑", width=3, command=lambda: self.move_segment_up(segment_num-1)).pack(side='left', padx=2)
        ttk.Button(btn_frame, text="↓", width=3, command=lambda: self.move_segment_down(segment_num-1)).pack(side='left', padx=2)
        ttk.Button(btn_frame, text="✕", width=3, command=lambda: self.remove_segment(segment_num-1)).pack(side='left', padx=2)

        self.mission_segments.append(segment_data)
        self.update_mission_summary()

    def remove_segment(self, index):
        """Remove a mission segment"""
        if 0 <= index < len(self.mission_segments):
            segment = self.mission_segments.pop(index)
            segment['frame'].destroy()
            self.rebuild_mission_display()
            self.update_mission_summary()

    def move_segment_up(self, index):
        """Move segment up in the list"""
        if index > 0:
            self.mission_segments[index], self.mission_segments[index-1] = \
                self.mission_segments[index-1], self.mission_segments[index]
            self.rebuild_mission_display()

    def move_segment_down(self, index):
        """Move segment down in the list"""
        if index < len(self.mission_segments) - 1:
            self.mission_segments[index], self.mission_segments[index+1] = \
                self.mission_segments[index+1], self.mission_segments[index]
            self.rebuild_mission_display()

    def rebuild_mission_display(self):
        """Rebuild the mission display after reordering"""
        # Clear all widgets
        for widget in self.segments_frame_inner.winfo_children():
            widget.destroy()

        # Rebuild from mission_segments list
        temp_segments = self.mission_segments.copy()
        self.mission_segments = []

        for seg in temp_segments:
            # Re-add each segment
            segment_type = seg['type'].replace('_', ' ').title()
            self.segment_type_var.set(segment_type)

            # Set parameters before adding
            if 'duration_s' in seg and hasattr(seg['duration_s'], 'get'):
                duration_val = seg['duration_s'].get()
            else:
                duration_val = "60"

            if 'speed_ms' in seg and hasattr(seg['speed_ms'], 'get'):
                speed_val = seg['speed_ms'].get()
            else:
                speed_val = "15.0"

            # Add segment (this will create new frame and variables)
            self.add_mission_segment()

            # Update with saved values
            if 'duration_s' in seg:
                self.mission_segments[-1]['duration_s'].set(duration_val)
            if 'speed_ms' in seg:
                self.mission_segments[-1]['speed_ms'].set(speed_val)

        self.update_mission_summary()

    def clear_mission_segments(self, confirm=True):
        """Clear all mission segments"""
        if confirm:
            if not messagebox.askyesno("Clear Mission", "Remove all segments?"):
                return

        for seg in self.mission_segments:
            seg['frame'].destroy()
        self.mission_segments = []
        self.update_mission_summary()

    def simulate_mission(self):
        """Simulate the mission and calculate energy"""
        if not self.current_calc:
            messagebox.showwarning("No Analysis", "Run an analysis first!")
            return

        if not self.mission_segments:
            messagebox.showwarning("Empty Mission", "Add some segments first!")
            return

        try:
            # Convert GUI segments to analysis format
            mission_profile = []
            for seg in self.mission_segments:
                seg_type = seg['type']

                if seg_type == "hover":
                    duration = float(seg['duration_s'].get())
                    mission_profile.append({
                        'type': 'hover',
                        'duration_s': duration
                    })
                elif seg_type == "cruise":
                    duration = float(seg['duration_s'].get())
                    speed = float(seg['speed_ms'].get())
                    mission_profile.append({
                        'type': 'cruise',
                        'duration_s': duration,
                        'speed_ms': speed
                    })
                elif seg_type == "transition_forward":
                    mission_profile.append({'type': 'transition_forward'})
                elif seg_type == "transition_back":
                    mission_profile.append({'type': 'transition_back'})

            # Analyze mission
            mission_results = self.current_calc.mission_profile_analysis(mission_profile)

            # Display results
            self.display_mission_results(mission_results)

        except Exception as e:
            messagebox.showerror("Simulation Error", f"Could not simulate mission:\n{e}")

    def display_mission_results(self, results):
        """Display mission simulation results"""
        total_time = results['total_time_min']
        total_energy = results['total_energy_wh']
        battery_cap = results['battery_capacity_wh']
        remaining = results['battery_remaining_percent']

        # Update summary
        summary = f"""Mission Analysis Results:
  Total Time:        {total_time:.1f} min ({total_time/60:.2f} hours)
  Total Energy:      {total_energy:.1f} Wh
  Battery Capacity:  {battery_cap:.1f} Wh
  Energy Remaining:  {results['energy_remaining_wh']:.1f} Wh ({remaining:.1f}%)

  Status: {'✓ FEASIBLE' if remaining >= 20 else '✗ NOT FEASIBLE (need >20% reserve)'}"""

        # Show in dialog
        messagebox.showinfo("Mission Simulation Results", summary)

        # Update summary text
        self.mission_summary_text.set(summary)

    def update_mission_summary(self):
        """Update mission summary display"""
        if not self.mission_segments:
            self.mission_summary_text.set("No segments added yet")
            return

        total_segments = len(self.mission_segments)
        hover_count = sum(1 for s in self.mission_segments if s['type'] == 'hover')
        cruise_count = sum(1 for s in self.mission_segments if s['type'] == 'cruise')
        trans_count = sum(1 for s in self.mission_segments if 'transition' in s['type'])

        summary = f"Segments: {total_segments} total ({hover_count} hover, {cruise_count} cruise, {trans_count} transitions)\nClick 'Simulate Mission' to analyze"
        self.mission_summary_text.set(summary)

    def save_mission(self):
        """Save mission profile to file"""
        messagebox.showinfo("Coming Soon", "Save mission profile to JSON file")

    def load_mission(self):
        """Load mission profile from file"""
        messagebox.showinfo("Coming Soon", "Load mission profile from JSON file")

    def load_mission_template(self, template_id=None):
        """Load a mission template"""
        try:
            from mission_templates import get_mission_template

            # Get template ID from dropdown if not provided
            if template_id is None:
                template_name = self.template_var.get()
                template_id = self.template_map.get(template_name)

            if not template_id:
                return

            template = get_mission_template(template_id)
            if not template:
                messagebox.showerror("Error", f"Template '{template_id}' not found")
                return

            # Clear existing segments
            self.clear_mission_segments(confirm=False)

            # Load template segments
            for seg in template['segments']:
                seg_type = seg['type']

                if seg_type == "hover":
                    self.segment_type_var.set("Hover")
                    self.add_mission_segment()
                    # Set duration
                    self.mission_segments[-1]['duration_s'].set(str(seg['duration_s']))

                elif seg_type == "cruise":
                    self.segment_type_var.set("Cruise")
                    self.add_mission_segment()
                    # Set duration and speed
                    self.mission_segments[-1]['duration_s'].set(str(seg['duration_s']))
                    self.mission_segments[-1]['speed_ms'].set(str(seg['speed_ms']))

                elif seg_type == "transition_forward":
                    self.segment_type_var.set("Transition Forward")
                    self.add_mission_segment()

                elif seg_type == "transition_back":
                    self.segment_type_var.set("Transition Back")
                    self.add_mission_segment()

            # Update summary
            self.update_mission_summary()

            # Show info
            est_time = template.get('estimated_time_min', 'N/A')
            est_dist = template.get('estimated_distance_km', 'N/A')
            self.update_status(
                f"✓ Loaded template: {template['name']} "
                f"(~{est_time} min, ~{est_dist} km)"
            )

        except Exception as e:
            messagebox.showerror("Load Template Error", f"Could not load template:\n{e}")

    # -----------------------------------------------------------------------
    # TAB 5: COMPARISON
    # -----------------------------------------------------------------------

    def create_comparison_tab(self):
        """Create multi-preset comparison tab"""
        # Top toolbar
        toolbar = ttk.Frame(self.tab_comparison)
        toolbar.pack(fill='x', padx=10, pady=5)

        ttk.Label(toolbar, text="Multi-Preset Comparison", style='Title.TLabel').pack(side='left')
        ttk.Button(toolbar, text="Run Comparison", command=self.run_comparison, style='Primary.TButton').pack(side='right', padx=5)
        ttk.Button(toolbar, text="Export Table", command=self.export_comparison_table).pack(side='right', padx=5)

        # Preset selection
        selection_frame = ttk.LabelFrame(self.tab_comparison, text=" Select Presets to Compare ", padding=10)
        selection_frame.pack(fill='x', padx=10, pady=5)

        self.comparison_presets = {}
        presets = self.preset_manager.list_presets()

        for preset_name in presets:
            var = tk.BooleanVar(value=True)
            desc = self.preset_manager.get_preset_description(preset_name)
            ttk.Checkbutton(selection_frame, text=desc, variable=var).pack(anchor='w', pady=2)
            self.comparison_presets[preset_name] = var

        # Results area
        results_frame = ttk.LabelFrame(self.tab_comparison, text=" Comparison Results ", padding=10)
        results_frame.pack(fill='both', expand=True, padx=10, pady=5)

        # Scrollable text for comparison table
        self.comparison_text = scrolledtext.ScrolledText(
            results_frame,
            wrap=tk.WORD,
            width=120,
            height=25,
            font=('Courier', 9)
        )
        self.comparison_text.pack(fill='both', expand=True)

        # Initial message
        self.comparison_text.insert('1.0', "Select presets above and click 'Run Comparison'")
        self.comparison_text.config(state='disabled')

    def run_comparison(self):
        """Run comparison analysis for selected presets"""
        try:
            self.update_status("Running comparison...")

            # Get selected presets
            selected = [name for name, var in self.comparison_presets.items() if var.get()]

            if len(selected) < 2:
                messagebox.showwarning("Selection Required", "Select at least 2 presets to compare!")
                return

            # Analyze each preset
            results = {}
            for preset_name in selected:
                config = self.preset_manager.get_preset(preset_name)
                calc = PerformanceCalculator(config)
                perf = calc.generate_performance_summary()
                results[preset_name] = perf

            # Display comparison
            self.display_comparison_results(results)

            self.update_status("Comparison complete!")

        except Exception as e:
            messagebox.showerror("Comparison Error", f"Could not run comparison:\n{e}")
            self.update_status("Comparison failed")

    def display_comparison_results(self, results):
        """Display comparison results in table format"""
        self.comparison_text.config(state='normal')
        self.comparison_text.delete('1.0', tk.END)

        output = []
        output.append("="*120)
        output.append(" MULTI-PRESET COMPARISON".center(120))
        output.append("="*120)
        output.append("")

        # Header row
        preset_names = list(results.keys())
        header = f"{'Parameter':<35}"
        for name in preset_names:
            header += f"{name.upper():<25}"
        output.append(header)
        output.append("-"*120)

        # Weight & Geometry
        output.append("\nWEIGHT & GEOMETRY:")
        self.add_comparison_row(output, preset_names, results, "Total Weight (kg)", lambda r: r['weight']['total_kg'])
        self.add_comparison_row(output, preset_names, results, "Wing Loading (kg/m²)", lambda r: r['weight']['wing_loading_kgm2'])

        # Performance
        output.append("\nPERFORMANCE:")
        self.add_comparison_row(output, preset_names, results, "Hover Endurance (pure hover, min)", lambda r: r['hover']['endurance_min'])
        self.add_comparison_row(output, preset_names, results, "Forward Flight Endurance (min)", lambda r: r['cruise']['endurance_min'])
        self.add_comparison_row(output, preset_names, results, "Forward Flight Range (km)", lambda r: r['cruise']['range_km'])
        self.add_comparison_row(output, preset_names, results, "Forward Flight Speed (m/s)", lambda r: r['cruise']['speed_ms'])
        self.add_comparison_row(output, preset_names, results, "Max Range - optimized (km)", lambda r: r['best_range']['range_km'])

        # Power
        output.append("\nPOWER BUDGET:")
        self.add_comparison_row(output, preset_names, results, "Hover Power (W)", lambda r: r['hover']['power_w'])
        self.add_comparison_row(output, preset_names, results, "Forward Flight Power (W)", lambda r: r['cruise']['power_w'])
        if 'power_budget' in results[preset_names[0]]['cruise']:
            self.add_comparison_row(output, preset_names, results, "Control Power (W)",
                lambda r: r['cruise']['power_budget']['control_power_w'] if 'power_budget' in r['cruise'] else 0)

        # Transitions
        if 'transitions' in results[preset_names[0]]:
            output.append("\nTRANSITIONS:")
            self.add_comparison_row(output, preset_names, results, "Forward Transition (Wh)",
                lambda r: r['transitions']['forward']['energy_wh'] if 'transitions' in r else 0)
            self.add_comparison_row(output, preset_names, results, "Back Transition (Wh)",
                lambda r: r['transitions']['back']['energy_wh'] if 'transitions' in r else 0)

        # Aerodynamics
        output.append("\nAERODYNAMICS:")
        self.add_comparison_row(output, preset_names, results, "Max L/D Ratio", lambda r: r['aerodynamics']['max_ld_ratio'])
        self.add_comparison_row(output, preset_names, results, "CD0", lambda r: r['aerodynamics']['cd0'])

        output.append("\n" + "="*120)

        # Insert into text widget
        self.comparison_text.insert('1.0', '\n'.join(output))
        self.comparison_text.config(state='disabled')

        # Store for export
        self.comparison_results = results

    def add_comparison_row(self, output, preset_names, results, parameter_name, value_func):
        """Add a comparison row to the output"""
        row = f"  {parameter_name:<33}"
        for name in preset_names:
            try:
                value = value_func(results[name])
                row += f"{value:<25.2f}"
            except:
                row += f"{'N/A':<25}"
        output.append(row)

    def export_comparison_table(self):
        """Export comparison table to CSV"""
        if not hasattr(self, 'comparison_results'):
            messagebox.showwarning("No Results", "Run a comparison first!")
            return

        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
            )

            if filename:
                import csv
                with open(filename, 'w', newline='') as f:
                    writer = csv.writer(f)

                    # Header
                    preset_names = list(self.comparison_results.keys())
                    header = ["Parameter"] + [name.upper() for name in preset_names]
                    writer.writerow(header)

                    # Data rows
                    rows = [
                        ("Hover Endurance - pure hover (min)", lambda r: r['hover']['endurance_min']),
                        ("Forward Flight Endurance (min)", lambda r: r['cruise']['endurance_min']),
                        ("Forward Flight Range (km)", lambda r: r['cruise']['range_km']),
                        ("Hover Power (W)", lambda r: r['hover']['power_w']),
                        ("Forward Flight Power (W)", lambda r: r['cruise']['power_w']),
                    ]

                    for param_name, value_func in rows:
                        row = [param_name]
                        for name in preset_names:
                            try:
                                row.append(f"{value_func(self.comparison_results[name]):.2f}")
                            except:
                                row.append("N/A")
                        writer.writerow(row)

                self.update_status(f"Comparison saved to {filename}")
                messagebox.showinfo("Success", f"Comparison saved to:\n{filename}")

        except Exception as e:
            messagebox.showerror("Export Error", f"Could not export comparison:\n{e}")

    # -----------------------------------------------------------------------
    # TAB 6: EXPORT MANAGER
    # -----------------------------------------------------------------------

    def create_export_tab(self):
        """Create export manager tab with PDF/Excel/CSV export capabilities"""
        # Top toolbar
        toolbar = ttk.Frame(self.tab_export)
        toolbar.pack(fill='x', padx=10, pady=5)

        ttk.Label(toolbar, text="Export Manager", style='Title.TLabel').pack(side='left')

        # Export format selection
        format_frame = ttk.LabelFrame(self.tab_export, text=" Export Format ", padding=10)
        format_frame.pack(fill='x', padx=10, pady=5)

        self.export_format_var = tk.StringVar(value="PDF Report")
        export_formats = [
            ("PDF Report", "PDF"),
            ("Excel Spreadsheet", "XLSX"),
            ("CSV Data", "CSV"),
            ("JSON Data", "JSON"),
            ("HTML Report", "HTML")
        ]

        for i, (label, value) in enumerate(export_formats):
            ttk.Radiobutton(
                format_frame,
                text=label,
                variable=self.export_format_var,
                value=value,
                command=self.on_export_format_changed
            ).grid(row=i//3, column=i%3, sticky='w', padx=20, pady=5)

        # Report template selection (for PDF/HTML)
        template_frame = ttk.LabelFrame(self.tab_export, text=" Report Template ", padding=10)
        template_frame.pack(fill='x', padx=10, pady=5)

        self.report_template_var = tk.StringVar(value="Engineering")
        templates = [
            ("Engineering Report (Full Details)", "Engineering"),
            ("Executive Summary (Key Results Only)", "Executive"),
            ("Flight Test Report (Field Ready)", "FlightTest"),
            ("Comparison Report (Multi-Preset)", "Comparison")
        ]

        for i, (label, value) in enumerate(templates):
            ttk.Radiobutton(
                template_frame,
                text=label,
                variable=self.report_template_var,
                value=value
            ).grid(row=i, column=0, sticky='w', padx=20, pady=3)

        # Export options
        options_frame = ttk.LabelFrame(self.tab_export, text=" Export Options ", padding=10)
        options_frame.pack(fill='x', padx=10, pady=5)

        self.include_plots_var = tk.BooleanVar(value=True)
        self.include_mission_var = tk.BooleanVar(value=True)
        self.include_comparison_var = tk.BooleanVar(value=False)
        self.open_after_export_var = tk.BooleanVar(value=True)

        ttk.Checkbutton(options_frame, text="Include plots and charts", variable=self.include_plots_var).grid(row=0, column=0, sticky='w', padx=10, pady=3)
        ttk.Checkbutton(options_frame, text="Include mission analysis", variable=self.include_mission_var).grid(row=1, column=0, sticky='w', padx=10, pady=3)
        ttk.Checkbutton(options_frame, text="Include preset comparisons", variable=self.include_comparison_var).grid(row=2, column=0, sticky='w', padx=10, pady=3)
        ttk.Checkbutton(options_frame, text="Open file after export", variable=self.open_after_export_var).grid(row=3, column=0, sticky='w', padx=10, pady=3)

        # Output directory
        output_frame = ttk.LabelFrame(self.tab_export, text=" Output Location ", padding=10)
        output_frame.pack(fill='x', padx=10, pady=5)

        ttk.Label(output_frame, text="Directory:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.output_dir_var = tk.StringVar(value="./output")
        ttk.Entry(output_frame, textvariable=self.output_dir_var, width=60).grid(row=0, column=1, sticky='ew', padx=5)
        ttk.Button(output_frame, text="Browse...", command=self.browse_output_directory).grid(row=0, column=2, padx=5)

        output_frame.columnconfigure(1, weight=1)

        # File preview
        preview_frame = ttk.LabelFrame(self.tab_export, text=" Export Preview ", padding=10)
        preview_frame.pack(fill='both', expand=True, padx=10, pady=5)

        self.export_preview_text = scrolledtext.ScrolledText(
            preview_frame,
            wrap=tk.WORD,
            width=100,
            height=15,
            font=('Courier', 9)
        )
        self.export_preview_text.pack(fill='both', expand=True)
        self.export_preview_text.insert('1.0', "Export preview will appear here...\n\nSelect format and click 'Generate Preview' to see what will be exported.")
        self.export_preview_text.config(state='disabled')

        # Action buttons
        action_frame = ttk.Frame(self.tab_export)
        action_frame.pack(fill='x', padx=10, pady=10)

        ttk.Button(action_frame, text="Generate Preview", command=self.generate_export_preview).pack(side='left', padx=5)
        ttk.Button(action_frame, text="Export", command=self.perform_export, style='Primary.TButton').pack(side='left', padx=5)
        ttk.Button(action_frame, text="Export All Formats", command=self.export_all_formats).pack(side='left', padx=5)
        ttk.Button(action_frame, text="Clear Preview", command=self.clear_export_preview).pack(side='left', padx=5)

    def on_export_format_changed(self):
        """Handle export format change"""
        format_type = self.export_format_var.get()
        self.update_status(f"Export format: {format_type}")

    def browse_output_directory(self):
        """Browse for output directory"""
        directory = filedialog.askdirectory(initialdir=self.output_dir_var.get())
        if directory:
            self.output_dir_var.set(directory)
            self.update_status(f"Output directory: {directory}")

    def generate_export_preview(self):
        """Generate preview of what will be exported"""
        if not self.current_results:
            messagebox.showwarning("No Results", "Run an analysis first!")
            return

        try:
            self.update_status("Generating preview...")

            format_type = self.export_format_var.get()
            template = self.report_template_var.get()

            # Clear preview
            self.export_preview_text.config(state='normal')
            self.export_preview_text.delete('1.0', tk.END)

            # Generate preview based on format
            if format_type == "PDF":
                preview = self.generate_pdf_preview(template)
            elif format_type == "XLSX":
                preview = self.generate_excel_preview()
            elif format_type == "CSV":
                preview = self.generate_csv_preview()
            elif format_type == "JSON":
                preview = self.generate_json_preview()
            elif format_type == "HTML":
                preview = self.generate_html_preview(template)
            else:
                preview = "Preview not available for this format"

            self.export_preview_text.insert('1.0', preview)
            self.export_preview_text.config(state='disabled')

            self.update_status("Preview generated")

        except Exception as e:
            messagebox.showerror("Preview Error", f"Could not generate preview:\n{e}")
            self.update_status("Preview failed")

    def generate_pdf_preview(self, template):
        """Generate PDF export preview"""
        preview = []
        preview.append("="*80)
        preview.append(f" PDF EXPORT PREVIEW - {template.upper()} TEMPLATE".center(80))
        preview.append("="*80)
        preview.append("")

        if template == "Engineering":
            preview.append("DOCUMENT STRUCTURE:")
            preview.append("  • Title Page")
            preview.append("  • Table of Contents")
            preview.append("  • Executive Summary")
            preview.append("  • Aircraft Configuration")
            preview.append("  • Performance Analysis")
            preview.append("    - Hover Performance")
            preview.append("    - Cruise Performance")
            preview.append("    - Power Budget Breakdown")
            preview.append("    - Transition Analysis")
            preview.append("  • Aerodynamic Analysis")
            preview.append("  • Charts and Graphs")
            if self.include_plots_var.get():
                preview.append("  • Custom Plots")
            if self.include_mission_var.get():
                preview.append("  • Mission Profile Analysis")
            preview.append("  • Conclusions and Recommendations")
            preview.append("")

        elif template == "Executive":
            preview.append("DOCUMENT STRUCTURE:")
            preview.append("  • Title Page")
            preview.append("  • Key Performance Metrics")
            preview.append("  • Performance Summary Table")
            preview.append("  • Critical Charts")
            preview.append("  • Recommendations")
            preview.append("")

        elif template == "FlightTest":
            preview.append("DOCUMENT STRUCTURE:")
            preview.append("  • Test Information")
            preview.append("  • Aircraft Configuration")
            preview.append("  • Predicted Performance")
            preview.append("  • Test Data Sheets (blank)")
            preview.append("  • Safety Notes")
            preview.append("")

        elif template == "Comparison":
            preview.append("DOCUMENT STRUCTURE:")
            preview.append("  • Title Page")
            preview.append("  • Comparison Overview")
            preview.append("  • Side-by-Side Performance Table")
            preview.append("  • Comparison Charts")
            preview.append("  • Best Use Cases")
            preview.append("")

        # Sample content
        preview.append("SAMPLE CONTENT:")
        preview.append("-"*80)
        preview.append(f"Preset: {self.current_preset_name.upper()}")
        preview.append(f"Weight: {self.current_config.total_takeoff_weight_kg:.2f} kg")
        preview.append(f"Hover Endurance (pure hover): {self.current_results['hover']['endurance_min']:.1f} min")
        preview.append(f"Forward Flight Range: {self.current_results['cruise']['range_km']:.1f} km")
        preview.append(f"Forward Flight Power: {self.current_results['cruise']['power_w']:.0f} W")
        preview.append("")

        preview.append("PDF FEATURES:")
        preview.append("  • Professional formatting with headers/footers")
        preview.append("  • Page numbers and table of contents")
        preview.append("  • High-resolution charts and graphs")
        preview.append("  • Color-coded sections")
        preview.append("  • Ready for printing")
        preview.append("")
        preview.append("="*80)

        return '\n'.join(preview)

    def generate_excel_preview(self):
        """Generate Excel export preview"""
        preview = []
        preview.append("="*80)
        preview.append(" EXCEL EXPORT PREVIEW".center(80))
        preview.append("="*80)
        preview.append("")

        preview.append("WORKBOOK STRUCTURE:")
        preview.append("")
        preview.append("Sheet 1: SUMMARY")
        preview.append("  ├── Key Performance Metrics")
        preview.append("  ├── Configuration Parameters")
        preview.append("  └── Quick Reference Table")
        preview.append("")

        preview.append("Sheet 2: HOVER PERFORMANCE")
        preview.append("  ├── Endurance: {:.1f} min".format(self.current_results['hover']['endurance_min']))
        preview.append("  ├── Power: {:.0f} W".format(self.current_results['hover']['power_w']))
        preview.append("  └── Current: {:.1f} A".format(self.current_results['hover']['current_a']))
        preview.append("")

        preview.append("Sheet 3: CRUISE PERFORMANCE")
        preview.append("  ├── Range: {:.1f} km".format(self.current_results['cruise']['range_km']))
        preview.append("  ├── Endurance: {:.1f} min".format(self.current_results['cruise']['endurance_min']))
        preview.append("  ├── Power: {:.0f} W".format(self.current_results['cruise']['power_w']))
        preview.append("  └── Speed: {:.1f} m/s".format(self.current_results['cruise']['speed_ms']))
        preview.append("")

        preview.append("Sheet 4: POWER BUDGET")
        preview.append("  └── Detailed breakdown by component")
        preview.append("")

        preview.append("Sheet 5: AERODYNAMICS")
        preview.append("  ├── L/D Ratio: {:.2f}".format(self.current_results['aerodynamics']['max_ld_ratio']))
        preview.append("  └── Drag components")
        preview.append("")

        if self.include_mission_var.get():
            preview.append("Sheet 6: MISSION ANALYSIS")
            preview.append("  └── Mission segment breakdown")
            preview.append("")

        preview.append("EXCEL FEATURES:")
        preview.append("  • Formatted tables with conditional formatting")
        preview.append("  • Embedded charts and graphs")
        preview.append("  • Formula-driven calculations")
        preview.append("  • Easy to import into other tools")
        preview.append("")
        preview.append("="*80)

        return '\n'.join(preview)

    def generate_csv_preview(self):
        """Generate CSV export preview"""
        preview = []
        preview.append("="*80)
        preview.append(" CSV EXPORT PREVIEW".center(80))
        preview.append("="*80)
        preview.append("")

        preview.append("CSV FORMAT: performance_data.csv")
        preview.append("")
        preview.append("Parameter,Value,Unit")
        preview.append(f"Preset,{self.current_preset_name},-")
        preview.append(f"Weight,{self.current_config.total_takeoff_weight_kg:.2f},kg")
        preview.append(f"Hover_Endurance,{self.current_results['hover']['endurance_min']:.1f},min")
        preview.append(f"Hover_Power,{self.current_results['hover']['power_w']:.0f},W")
        preview.append(f"Cruise_Range,{self.current_results['cruise']['range_km']:.1f},km")
        preview.append(f"Cruise_Endurance,{self.current_results['cruise']['endurance_min']:.1f},min")
        preview.append(f"Cruise_Power,{self.current_results['cruise']['power_w']:.0f},W")
        preview.append(f"Cruise_Speed,{self.current_results['cruise']['speed_ms']:.1f},m/s")
        preview.append(f"Max_LD_Ratio,{self.current_results['aerodynamics']['max_ld_ratio']:.2f},-")
        preview.append("...")
        preview.append("")

        preview.append("CSV FILES TO BE CREATED:")
        preview.append("  • performance_summary.csv - Key metrics")
        preview.append("  • configuration.csv - All parameters")
        preview.append("  • power_budget.csv - Detailed breakdown")
        if self.include_mission_var.get():
            preview.append("  • mission_analysis.csv - Mission segments")
        preview.append("")
        preview.append("="*80)

        return '\n'.join(preview)

    def generate_json_preview(self):
        """Generate JSON export preview"""
        import json

        preview = []
        preview.append("="*80)
        preview.append(" JSON EXPORT PREVIEW".center(80))
        preview.append("="*80)
        preview.append("")

        # Create simplified JSON structure
        json_data = {
            "preset": self.current_preset_name,
            "timestamp": "2025-01-20T12:00:00",
            "configuration": {
                "weight_kg": self.current_config.total_takeoff_weight_kg,
                "wingspan_m": self.current_config.wingspan_m,
                "wing_chord_m": self.current_config.wing_chord_m,
            },
            "performance": {
                "hover": {
                    "endurance_min": float(self.current_results['hover']['endurance_min']),
                    "power_w": float(self.current_results['hover']['power_w']),
                },
                "cruise": {
                    "range_km": float(self.current_results['cruise']['range_km']),
                    "endurance_min": float(self.current_results['cruise']['endurance_min']),
                    "power_w": float(self.current_results['cruise']['power_w']),
                    "speed_ms": float(self.current_results['cruise']['speed_ms']),
                }
            }
        }

        preview.append(json.dumps(json_data, indent=2))
        preview.append("")
        preview.append("JSON FEATURES:")
        preview.append("  • Complete data export")
        preview.append("  • Easy to parse programmatically")
        preview.append("  • Compatible with web APIs")
        preview.append("  • Human-readable format")
        preview.append("")
        preview.append("="*80)

        return '\n'.join(preview)

    def generate_html_preview(self, template):
        """Generate HTML export preview"""
        preview = []
        preview.append("="*80)
        preview.append(f" HTML EXPORT PREVIEW - {template.upper()} TEMPLATE".center(80))
        preview.append("="*80)
        preview.append("")

        preview.append("HTML STRUCTURE:")
        preview.append("<!DOCTYPE html>")
        preview.append("<html>")
        preview.append("  <head>")
        preview.append("    <title>VTOL Performance Analysis</title>")
        preview.append("    <style>/* Professional CSS styling */</style>")
        preview.append("  </head>")
        preview.append("  <body>")
        preview.append("    <header>")
        preview.append("      <h1>VTOL Performance Analysis Report</h1>")
        preview.append(f"      <h2>Preset: {self.current_preset_name.upper()}</h2>")
        preview.append("    </header>")
        preview.append("    <main>")
        preview.append("      <section id='summary'>")
        preview.append("        <!-- Key Performance Metrics -->")
        preview.append("      </section>")
        preview.append("      <section id='charts'>")
        preview.append("        <!-- Interactive Charts -->")
        preview.append("      </section>")
        preview.append("      <section id='details'>")
        preview.append("        <!-- Detailed Analysis -->")
        preview.append("      </section>")
        preview.append("    </main>")
        preview.append("  </body>")
        preview.append("</html>")
        preview.append("")

        preview.append("HTML FEATURES:")
        preview.append("  • Responsive design (mobile-friendly)")
        preview.append("  • Interactive charts with Chart.js")
        preview.append("  • Collapsible sections")
        preview.append("  • Print-optimized CSS")
        preview.append("  • Can be hosted on web servers")
        preview.append("")
        preview.append("="*80)

        return '\n'.join(preview)

    def clear_export_preview(self):
        """Clear export preview"""
        self.export_preview_text.config(state='normal')
        self.export_preview_text.delete('1.0', tk.END)
        self.export_preview_text.insert('1.0', "Export preview cleared.")
        self.export_preview_text.config(state='disabled')

    def perform_export(self):
        """Perform the actual export"""
        if not self.current_results:
            messagebox.showwarning("No Results", "Run an analysis first!")
            return

        try:
            self.update_status("Exporting...")

            format_type = self.export_format_var.get()
            output_dir = self.output_dir_var.get()

            # Create output directory
            import os
            os.makedirs(output_dir, exist_ok=True)

            # Export based on format
            if format_type == "PDF":
                filepath = self.export_pdf()
            elif format_type == "XLSX":
                filepath = self.export_excel()
            elif format_type == "CSV":
                filepath = self.export_csv()
            elif format_type == "JSON":
                filepath = self.export_json()
            elif format_type == "HTML":
                filepath = self.export_html()
            else:
                raise ValueError(f"Unknown format: {format_type}")

            self.update_status(f"Exported to {filepath}")
            messagebox.showinfo("Export Successful", f"File exported to:\n{filepath}")

            # Open file if requested
            if self.open_after_export_var.get() and filepath:
                self.open_file(filepath)

        except Exception as e:
            messagebox.showerror("Export Error", f"Could not export file:\n{e}")
            self.update_status("Export failed")

    def export_pdf(self):
        """Export analysis as PDF report"""
        try:
            # Try to import reportlab
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
            from reportlab.lib import colors
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        except ImportError:
            messagebox.showwarning("PDF Export",
                "reportlab library not installed.\n\n"
                "Install with: pip install reportlab\n\n"
                "Falling back to text export...")
            return self.export_text()

        output_dir = self.output_dir_var.get()
        template = self.report_template_var.get()
        filename = os.path.join(output_dir, f"vtol_analysis_{self.current_preset_name}_{template.lower()}.pdf")

        # Create PDF
        doc = SimpleDocTemplate(filename, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2C3E50'),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        story.append(Paragraph("VTOL Performance Analysis Report", title_style))
        story.append(Paragraph(f"Preset: {self.current_preset_name.upper()}", styles['Heading2']))
        story.append(Spacer(1, 0.3*inch))

        # Summary table
        summary_data = [
            ['Parameter', 'Value', 'Unit'],
            ['Hover Endurance (pure hover)', f"{self.current_results['hover']['endurance_min']:.1f}", 'min'],
            ['Forward Flight Range', f"{self.current_results['cruise']['range_km']:.1f}", 'km'],
            ['Forward Flight Power', f"{self.current_results['cruise']['power_w']:.0f}", 'W'],
            ['Max L/D Ratio', f"{self.current_results['aerodynamics']['max_ld_ratio']:.2f}", '-'],
        ]

        table = Table(summary_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        story.append(table)
        story.append(Spacer(1, 0.5*inch))

        # Configuration section
        if template in ["Engineering", "FlightTest"]:
            story.append(Paragraph("Aircraft Configuration", styles['Heading2']))
            config_data = [
                ['Parameter', 'Value'],
                ['Weight', f"{self.current_config.total_takeoff_weight_kg:.2f} kg"],
                ['Wing Span', f"{self.current_config.wingspan_m:.2f} m"],
                ['Wing Chord', f"{self.current_config.wing_chord_m:.3f} m"],
                ['Wing Area', f"{self.current_config.wing_area_m2:.3f} m²"],
            ]

            config_table = Table(config_data)
            config_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27AE60')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            story.append(config_table)
            story.append(Spacer(1, 0.3*inch))

        # Build PDF
        doc.build(story)

        return filename

    def export_text(self):
        """Export as plain text (fallback when PDF not available)"""
        output_dir = self.output_dir_var.get()
        filename = os.path.join(output_dir, f"vtol_analysis_{self.current_preset_name}.txt")

        with open(filename, 'w', encoding='utf-8') as f:
            f.write(self.results_text.get('1.0', tk.END))

        return filename

    def export_excel(self):
        """Export analysis as Excel spreadsheet"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showwarning("Excel Export",
                "openpyxl library not installed.\n\n"
                "Install with: pip install openpyxl\n\n"
                "Falling back to CSV export...")
            return self.export_csv()

        output_dir = self.output_dir_var.get()
        filename = os.path.join(output_dir, f"vtol_analysis_{self.current_preset_name}.xlsx")

        # Create workbook
        wb = openpyxl.Workbook()

        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Header
        ws_summary['A1'] = "VTOL Performance Analysis"
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary['A2'] = f"Preset: {self.current_preset_name.upper()}"
        ws_summary['A2'].font = Font(size=12, bold=True)

        # Performance summary
        row = 4
        ws_summary[f'A{row}'] = "Parameter"
        ws_summary[f'B{row}'] = "Value"
        ws_summary[f'C{row}'] = "Unit"

        # Style header
        for col in ['A', 'B', 'C']:
            cell = ws_summary[f'{col}{row}']
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        data = [
            ("Hover Endurance (pure hover)", self.current_results['hover']['endurance_min'], "min"),
            ("Hover Power", self.current_results['hover']['power_w'], "W"),
            ("Forward Flight Range", self.current_results['cruise']['range_km'], "km"),
            ("Forward Flight Endurance", self.current_results['cruise']['endurance_min'], "min"),
            ("Forward Flight Power", self.current_results['cruise']['power_w'], "W"),
            ("Forward Flight Speed", self.current_results['cruise']['speed_ms'], "m/s"),
            ("Max L/D Ratio", self.current_results['aerodynamics']['max_ld_ratio'], "-"),
        ]

        for i, (param, value, unit) in enumerate(data, start=row+1):
            ws_summary[f'A{i}'] = param
            ws_summary[f'B{i}'] = f"{value:.2f}"
            ws_summary[f'C{i}'] = unit

        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 10

        # Sheet 2: Configuration
        ws_config = wb.create_sheet("Configuration")
        ws_config['A1'] = "Aircraft Configuration"
        ws_config['A1'].font = Font(size=14, bold=True)

        config_data = [
            ("Total Weight", self.current_config.total_takeoff_weight_kg, "kg"),
            ("Wing Span", self.current_config.wingspan_m, "m"),
            ("Wing Chord", self.current_config.wing_chord_m, "m"),
            ("Wing Area", self.current_config.wing_area_m2, "m²"),
            ("Battery Capacity", self.current_config.battery_capacity_wh, "Wh"),
        ]

        row = 3
        for param, value, unit in config_data:
            ws_config[f'A{row}'] = param
            ws_config[f'B{row}'] = f"{value:.3f}"
            ws_config[f'C{row}'] = unit
            row += 1

        # Save workbook
        wb.save(filename)

        return filename

    def export_csv(self):
        """Export analysis as CSV files"""
        import csv

        output_dir = self.output_dir_var.get()
        base_filename = f"vtol_analysis_{self.current_preset_name}"

        # Main performance CSV
        filename = os.path.join(output_dir, f"{base_filename}_performance.csv")

        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Parameter', 'Value', 'Unit'])
            writer.writerow(['Preset', self.current_preset_name, '-'])
            writer.writerow(['Hover_Endurance', f"{self.current_results['hover']['endurance_min']:.1f}", 'min'])
            writer.writerow(['Hover_Power', f"{self.current_results['hover']['power_w']:.0f}", 'W'])
            writer.writerow(['Cruise_Range', f"{self.current_results['cruise']['range_km']:.1f}", 'km'])
            writer.writerow(['Cruise_Endurance', f"{self.current_results['cruise']['endurance_min']:.1f}", 'min'])
            writer.writerow(['Cruise_Power', f"{self.current_results['cruise']['power_w']:.0f}", 'W'])
            writer.writerow(['Cruise_Speed', f"{self.current_results['cruise']['speed_ms']:.1f}", 'm/s'])
            writer.writerow(['Max_LD_Ratio', f"{self.current_results['aerodynamics']['max_ld_ratio']:.2f}", '-'])

        return filename

    def export_json(self):
        """Export analysis as JSON file"""
        import json

        output_dir = self.output_dir_var.get()
        filename = os.path.join(output_dir, f"vtol_analysis_{self.current_preset_name}.json")

        # Convert results to JSON-serializable format
        def make_serializable(obj):
            if isinstance(obj, dict):
                return {k: make_serializable(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [make_serializable(item) for item in obj]
            elif hasattr(obj, 'item'):  # numpy types
                return obj.item()
            elif hasattr(obj, '__dict__'):
                return str(obj)
            else:
                return obj

        export_data = {
            "preset": self.current_preset_name,
            "configuration": {
                "weight_kg": self.current_config.total_takeoff_weight_kg,
                "wingspan_m": self.current_config.wingspan_m,
                "wing_chord_m": self.current_config.wing_chord_m,
                "wing_area_m2": self.current_config.wing_area_m2,
                "battery_capacity_wh": self.current_config.battery_capacity_wh,
            },
            "performance": make_serializable(self.current_results)
        }

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2)

        return filename

    def export_html(self):
        """Export analysis as HTML report"""
        output_dir = self.output_dir_var.get()
        template = self.report_template_var.get()
        filename = os.path.join(output_dir, f"vtol_analysis_{self.current_preset_name}_{template.lower()}.html")

        html = []
        html.append("<!DOCTYPE html>")
        html.append("<html>")
        html.append("<head>")
        html.append("    <meta charset='UTF-8'>")
        html.append("    <title>VTOL Performance Analysis</title>")
        html.append("    <style>")
        html.append("        body { font-family: Arial, sans-serif; margin: 40px; background: #f5f5f5; }")
        html.append("        .container { max-width: 1200px; margin: 0 auto; background: white; padding: 30px; box-shadow: 0 0 10px rgba(0,0,0,0.1); }")
        html.append("        h1 { color: #2C3E50; border-bottom: 3px solid #3498DB; padding-bottom: 10px; }")
        html.append("        h2 { color: #3498DB; margin-top: 30px; }")
        html.append("        table { width: 100%; border-collapse: collapse; margin: 20px 0; }")
        html.append("        th { background: #3498DB; color: white; padding: 12px; text-align: left; }")
        html.append("        td { padding: 10px; border: 1px solid #ddd; }")
        html.append("        tr:nth-child(even) { background: #f9f9f9; }")
        html.append("        .metric { font-size: 24px; font-weight: bold; color: #27AE60; }")
        html.append("    </style>")
        html.append("</head>")
        html.append("<body>")
        html.append("    <div class='container'>")
        html.append("        <h1>VTOL Performance Analysis Report</h1>")
        html.append(f"        <h2>Preset: {self.current_preset_name.upper()}</h2>")
        html.append("")
        html.append("        <h2>Key Performance Metrics</h2>")
        html.append("        <table>")
        html.append("            <tr><th>Parameter</th><th>Value</th><th>Unit</th></tr>")
        html.append(f"            <tr><td>Hover Endurance (pure hover)</td><td class='metric'>{self.current_results['hover']['endurance_min']:.1f}</td><td>min</td></tr>")
        html.append(f"            <tr><td>Forward Flight Range</td><td class='metric'>{self.current_results['cruise']['range_km']:.1f}</td><td>km</td></tr>")
        html.append(f"            <tr><td>Forward Flight Power</td><td class='metric'>{self.current_results['cruise']['power_w']:.0f}</td><td>W</td></tr>")
        html.append(f"            <tr><td>Max L/D Ratio</td><td class='metric'>{self.current_results['aerodynamics']['max_ld_ratio']:.2f}</td><td>-</td></tr>")
        html.append("        </table>")
        html.append("")
        html.append("        <h2>Aircraft Configuration</h2>")
        html.append("        <table>")
        html.append(f"            <tr><td>Total Weight</td><td>{self.current_config.total_takeoff_weight_kg:.2f} kg</td></tr>")
        html.append(f"            <tr><td>Wing Span</td><td>{self.current_config.wingspan_m:.2f} m</td></tr>")
        html.append(f"            <tr><td>Wing Area</td><td>{self.current_config.wing_area_m2:.3f} m²</td></tr>")
        html.append(f"            <tr><td>Battery Capacity</td><td>{self.current_config.battery_capacity_wh:.1f} Wh</td></tr>")
        html.append("        </table>")
        html.append("    </div>")
        html.append("</body>")
        html.append("</html>")

        with open(filename, 'w', encoding='utf-8') as f:
            f.write('\n'.join(html))

        return filename

    def export_all_formats(self):
        """Export in all available formats"""
        if not self.current_results:
            messagebox.showwarning("No Results", "Run an analysis first!")
            return

        if messagebox.askyesno("Export All", "Export report in all formats?\n(PDF, Excel, CSV, JSON, HTML)"):
            try:
                self.update_status("Exporting all formats...")

                exported = []
                for format_type in ["PDF", "XLSX", "CSV", "JSON", "HTML"]:
                    self.export_format_var.set(format_type)
                    try:
                        filepath = self.perform_single_export(format_type)
                        if filepath:
                            exported.append(f"  • {format_type}: {os.path.basename(filepath)}")
                    except Exception as e:
                        exported.append(f"  • {format_type}: FAILED ({str(e)[:30]}...)")

                messagebox.showinfo("Batch Export Complete",
                    "Exported files:\n\n" + "\n".join(exported))

                self.update_status("All formats exported")

            except Exception as e:
                messagebox.showerror("Batch Export Error", f"Error during batch export:\n{e}")

    def perform_single_export(self, format_type):
        """Perform single export without dialogs"""
        output_dir = self.output_dir_var.get()
        os.makedirs(output_dir, exist_ok=True)

        if format_type == "PDF":
            return self.export_pdf()
        elif format_type == "XLSX":
            return self.export_excel()
        elif format_type == "CSV":
            return self.export_csv()
        elif format_type == "JSON":
            return self.export_json()
        elif format_type == "HTML":
            return self.export_html()

    def open_file(self, filepath):
        """Open file with default system application"""
        import subprocess
        import platform

        try:
            if platform.system() == 'Darwin':  # macOS
                subprocess.call(('open', filepath))
            elif platform.system() == 'Windows':  # Windows
                os.startfile(filepath)
            else:  # Linux
                subprocess.call(('xdg-open', filepath))
        except:
            pass  # Silently fail if can't open

    # -----------------------------------------------------------------------
    # DESIGN SCHEMATIC TAB (v4.1)
    # -----------------------------------------------------------------------

    def create_schematic_tab(self):
        """Create design schematic visualization tab (v4.1)"""
        # Toolbar with title and update button
        toolbar = ttk.Frame(self.tab_schematic)
        toolbar.pack(fill='x', padx=10, pady=10)

        ttk.Label(
            toolbar,
            text="Aircraft Design Schematic",
            style='Title.TLabel'
        ).pack(side='left')

        # Separator
        ttk.Separator(toolbar, orient='vertical').pack(side='left', padx=20, fill='y')

        # Info label
        ttk.Label(
            toolbar,
            text="3-view engineering drawing based on current parameters",
            font=('Arial', 10)
        ).pack(side='left', padx=10)

        # Update button on the right
        ttk.Button(
            toolbar,
            text="🔄 Update Schematic",
            command=self.update_schematic,
            style='Primary.TButton'
        ).pack(side='right', padx=5)

        # Instructions panel
        instructions_frame = ttk.LabelFrame(
            self.tab_schematic,
            text=" Instructions ",
            padding=10
        )
        instructions_frame.pack(fill='x', padx=10, pady=(0, 10))

        instructions_text = (
            "This tab shows a professional 3-view engineering drawing of your aircraft design.\n\n"
            "• Top View: Shows wing, fuselage, tail fins, and propeller positions\n"
            "• Front View: Shows fuselage cross-section and tail fin arrangement\n"
            "• Side View: Shows profile with wing and tail fin airfoils\n\n"
            "Change parameters in the Configuration tab, then click 'Update Schematic' to refresh the visualization."
        )

        ttk.Label(
            instructions_frame,
            text=instructions_text,
            font=('Arial', 10),
            justify='left'
        ).pack(anchor='w')

        # Canvas frame for matplotlib figure
        self.schematic_canvas_frame = ttk.Frame(self.tab_schematic)
        self.schematic_canvas_frame.pack(fill='both', expand=True, padx=10, pady=10)

        # Placeholder message (shown until first update)
        self.schematic_placeholder = ttk.Label(
            self.schematic_canvas_frame,
            text="Click 'Update Schematic' to visualize your aircraft design\n\n"
                 "The schematic will show top, front, and side views with dimensions",
            font=('Arial', 11),
            justify='center',
            foreground='#7f8c8d'
        )
        self.schematic_placeholder.pack(expand=True)

    def update_schematic(self):
        """Update schematic drawing based on current parameters (v4.1)"""
        try:
            self.update_status("Generating schematic...")

            # Clear previous content
            for widget in self.schematic_canvas_frame.winfo_children():
                widget.destroy()

            # Get current configuration from UI
            config = self.get_current_config()

            # Import schematic drawer
            from schematic import DroneSchematicDrawer

            # Create drawer and generate 3-view figure
            drawer = DroneSchematicDrawer(config)
            fig = drawer.draw_3_view(figsize=(15, 5))

            # Embed matplotlib figure in tkinter
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            canvas = FigureCanvasTkAgg(fig, master=self.schematic_canvas_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True)

            self.update_status("✓ Schematic updated successfully")

        except ImportError as e:
            messagebox.showerror(
                "Import Error",
                f"Could not import schematic drawer module:\n{str(e)}\n\n"
                "Ensure drone_schematic_drawer.py is in the same directory."
            )
            self.update_status("✗ Schematic generation failed")
        except Exception as e:
            messagebox.showerror(
                "Schematic Error",
                f"Error generating schematic:\n{str(e)}\n\n"
                "Check your parameter values and try again."
            )
            self.update_status("✗ Schematic generation failed")

    # -----------------------------------------------------------------------
    # STATUS BAR
    # -----------------------------------------------------------------------

    def create_status_bar(self):
        """Create bottom status bar"""
        self.status_bar = ttk.Frame(self, relief='sunken')
        self.status_bar.pack(side='bottom', fill='x')

        self.status_text = tk.StringVar(value="Ready")
        ttk.Label(self.status_bar, textvariable=self.status_text, style='Status.TLabel').pack(side='left', padx=10)

        self.preset_status = tk.StringVar(value="Preset: BASELINE")
        ttk.Label(self.status_bar, textvariable=self.preset_status, style='Status.TLabel').pack(side='right', padx=10)

    # -----------------------------------------------------------------------
    # CORE FUNCTIONS
    # -----------------------------------------------------------------------

    def load_preset(self, preset_name):
        """Load a preset configuration"""
        try:
            self.current_config = self.preset_manager.get_preset(preset_name)
            self.current_preset_name = preset_name
            self.update_ui_from_config()
            self.update_status(f"Loaded preset: {preset_name}")
            self.preset_status.set(f"Preset: {preset_name.upper()}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not load preset: {e}")

    def update_ui_from_config(self):
        """Update UI widgets from current configuration"""
        if not self.current_config:
            return

        # Map of parameter names to widget keys
        param_map = {
            'total_takeoff_weight_kg': 'total_takeoff_weight_kg',
            'wingspan_m': 'wingspan_m',
            'wing_chord_m': 'wing_chord_m',
            'field_elevation_m': 'field_elevation_m',
            'control_power_base_w': 'control_power_base_w',
            'control_power_speed_factor': 'control_power_speed_factor',
            'cd0_motor_nacelles': 'cd0_motor_nacelles',
            'cd0_fuselage_base': 'cd0_fuselage_base',
            'cd0_landing_gear': 'cd0_landing_gear',
            'cd0_interference': 'cd0_interference',
            'transition_forward_duration_s': 'transition_forward_duration_s',
            'transition_forward_power_factor': 'transition_forward_power_factor',
            'transition_back_duration_s': 'transition_back_duration_s',
            'transition_back_power_factor': 'transition_back_power_factor',
            'prop_efficiency_lowspeed': 'prop_efficiency_lowspeed',
            'prop_efficiency_highspeed': 'prop_efficiency_highspeed',
            'motor_efficiency_peak': 'motor_efficiency_peak',
            'esc_efficiency': 'esc_efficiency',
            'avionics_power_w': 'avionics_power_w',
            'payload_power_w': 'payload_power_w',
            'heater_power_w': 'heater_power_w',
            # Geometry parameters (v4.1)
            'fuselage_length_m': 'fuselage_length_m',
            'fuselage_diameter_m': 'fuselage_diameter_m',
            'num_tail_fins': 'num_tail_fins',
            'tail_fin_chord_m': 'tail_fin_chord_m',
            'tail_fin_span_m': 'tail_fin_span_m',
            'tail_fin_position_m': 'tail_fin_position_m',
            'tail_fin_thickness_ratio': 'tail_fin_thickness_ratio',
            'tail_fin_taper_ratio': 'tail_fin_taper_ratio',
            'motor_spacing_m': 'motor_spacing_m',
            'num_motors': 'num_motors',
        }

        for config_param, widget_key in param_map.items():
            if widget_key in self.param_widgets:
                widget = self.param_widgets[widget_key]
                value = getattr(self.current_config, config_param, '')

                # Convert percentages
                if 'efficiency' in config_param:
                    value = value * 100  # Show as percentage

                widget.delete(0, tk.END)
                widget.insert(0, str(value))

    def run_analysis(self):
        """Run performance analysis with current configuration"""
        try:
            # First apply any UI changes
            self.update_config_from_ui()

            # Validate configuration
            self.update_status("Validating configuration...")
            if not self.validate_config_silent():
                response = messagebox.askyesno(
                    "Validation Warning",
                    "Configuration has warnings. Continue with analysis anyway?"
                )
                if not response:
                    self.update_status("Analysis cancelled")
                    return

            self.update_status("Running analysis... Please wait")
            self.update()  # Force UI update

            # Create calculator
            self.current_calc = PerformanceCalculator(self.current_config)

            # Generate results
            self.current_results = self.current_calc.generate_performance_summary()

            # Display results
            self.display_results()

            self.update_status(f"✓ Analysis complete! Preset: {self.current_preset_name.upper()}")

            # Auto-switch to results tab
            self.notebook.select(1)

        except ValueError as e:
            messagebox.showerror("Invalid Input",
                f"Invalid parameter value:\n\n{e}\n\n"
                "Please check your inputs and try again.")
            self.update_status("Analysis failed - invalid input")
        except Exception as e:
            messagebox.showerror("Analysis Error",
                f"Could not run analysis:\n\n{e}\n\n"
                "Please check your configuration and try again.")
            self.update_status("Analysis failed")
            import traceback
            traceback.print_exc()

    def validate_config_silent(self):
        """Validate configuration without showing messages"""
        try:
            errors = []

            # Basic validation
            if self.current_config.total_takeoff_weight_kg < 1.0 or self.current_config.total_takeoff_weight_kg > 20.0:
                errors.append("Total weight out of range")

            if self.current_config.wingspan_m < 0.5 or self.current_config.wingspan_m > 5.0:
                errors.append("Wing span out of range")

            if self.current_config.prop_efficiency_lowspeed < 0.3 or self.current_config.prop_efficiency_lowspeed > 1.0:
                errors.append("Propeller efficiency out of range")

            return len(errors) == 0

        except:
            return False

    def display_results(self):
        """Display analysis results in results tab"""
        if not self.current_results:
            return

        # Clear previous results
        self.results_text.config(state='normal')
        self.results_text.delete('1.0', tk.END)

        # Format results (simplified version of v3.0 output)
        output = []
        output.append("="*80)
        output.append(" VTOL QUADPLANE PERFORMANCE ANALYSIS v4.0 - GUI".center(80))
        output.append("="*80)
        output.append("")

        # Key performance
        output.append("-"*80)
        output.append("KEY PERFORMANCE")
        output.append("-"*80)
        hover = self.current_results['hover']
        cruise = self.current_results['cruise']
        best_range = self.current_results['best_range']

        output.append(f"  Hover Endurance (pure hover):     {hover['endurance_min']:.1f} min")
        output.append(f"  Forward Flight Endurance:         {cruise['endurance_min']:.1f} min")
        output.append(f"  Forward Flight Range:             {cruise['range_km']:.1f} km @ {cruise['speed_ms']:.1f} m/s")
        output.append(f"  Forward Flight Power:             {cruise['power_w']:.0f} W")
        output.append(f"  Max Range (optimized speed):      {best_range['range_km']:.1f} km")
        output.append("")

        # Power budget
        if 'power_budget' in cruise:
            pb = cruise['power_budget']
            output.append("-"*80)
            output.append("FORWARD FLIGHT POWER BUDGET")
            output.append("-"*80)
            output.append(f"  Aerodynamic Drag:      {pb['aerodynamic_drag_w']:6.1f} W")
            output.append(f"  Propeller Efficiency:  {pb['propeller_efficiency']*100:6.1f} %")
            output.append(f"  Motor Electrical:      {pb['motor_electrical_w']:6.1f} W")
            output.append(f"  Control Power:         {pb['control_power_w']:6.1f} W")
            output.append(f"  Avionics:              {pb['avionics_w']:6.1f} W")
            output.append(f"  Payload:               {pb['payload_w']:6.1f} W")
            output.append(f"  ESC Loss:              {pb['esc_loss_w']:6.1f} W")
            output.append(f"  " + "-"*30)
            output.append(f"  TOTAL:                 {pb['total_electrical_w']:6.1f} W")
            output.append("")

        # Transitions
        if 'transitions' in self.current_results:
            trans = self.current_results['transitions']
            output.append("-"*80)
            output.append("TRANSITIONS")
            output.append("-"*80)
            output.append(f"  Forward: {trans['forward']['energy_wh']:.1f} Wh | Back: {trans['back']['energy_wh']:.1f} Wh")
            output.append(f"  Total Cycle: {trans['forward']['energy_wh'] + trans['back']['energy_wh']:.1f} Wh")
            output.append("")

        output.append("="*80)

        # Insert into text widget
        self.results_text.insert('1.0', '\n'.join(output))
        self.results_text.config(state='disabled')

    def update_status(self, message):
        """Update status bar message"""
        self.status_text.set(message)
        self.update_idletasks()

    # -----------------------------------------------------------------------
    # MENU ACTIONS
    # -----------------------------------------------------------------------

    def new_analysis(self):
        """Start new analysis"""
        if self.config_modified:
            response = messagebox.askyesnocancel(
                "Unsaved Changes",
                "Save current configuration before starting new analysis?"
            )
            if response is None:  # Cancel
                return
            elif response:  # Yes, save
                self.save_config()

        if messagebox.askyesno("New Analysis", "Reset to default configuration?"):
            self.load_preset("baseline")
            self.config_modified = False

    def open_config(self):
        """Open configuration from file"""
        try:
            filename = filedialog.askopenfilename(
                title="Open Configuration",
                initialdir=self.config_dir,
                filetypes=[
                    ("JSON files", "*.json"),
                    ("All files", "*.*")
                ]
            )

            if filename:
                self.load_config_from_file(filename)
                self.add_to_recent(filename)
                self.config_modified = False
                self.update_status(f"Loaded configuration from {os.path.basename(filename)}")

        except Exception as e:
            messagebox.showerror("Load Error", f"Could not load configuration:\n{e}")

    def save_config(self):
        """Save current configuration"""
        try:
            filename = filedialog.asksaveasfilename(
                title="Save Configuration",
                initialdir=self.config_dir,
                defaultextension=".json",
                filetypes=[
                    ("JSON files", "*.json"),
                    ("All files", "*.*")
                ]
            )

            if filename:
                self.save_config_to_file(filename)
                self.add_to_recent(filename)
                self.config_modified = False
                self.update_status(f"Configuration saved to {os.path.basename(filename)}")

        except Exception as e:
            messagebox.showerror("Save Error", f"Could not save configuration:\n{e}")

    def export_report(self):
        """Export analysis report"""
        messagebox.showinfo("Coming Soon", "Export PDF report")

    def export_all_data(self):
        """Export all data"""
        messagebox.showinfo("Coming Soon", "Export all data (CSV, JSON)")

    def toggle_auto_update(self):
        """Toggle auto-update mode"""
        if self.auto_update.get():
            self.update_status("Auto-update enabled")
        else:
            self.update_status("Auto-update disabled")

    def toggle_auto_save(self):
        """Toggle auto-save mode"""
        if self.auto_save_enabled.get():
            self.start_auto_save_timer()
            self.update_status("Auto-save enabled (saves every 5 minutes)")
        else:
            if self.auto_save_timer_id:
                self.after_cancel(self.auto_save_timer_id)
                self.auto_save_timer_id = None
            self.update_status("Auto-save disabled")

    def toggle_dark_mode(self):
        """Toggle dark mode"""
        messagebox.showinfo("Coming Soon", "Dark mode toggle")

    def refresh_results(self):
        """Refresh analysis results"""
        if self.current_results:
            self.display_results()
            self.update_status("Results refreshed")
        else:
            messagebox.showinfo("No Results", "Run an analysis first")

    def optimize_parameters(self):
        """Open parameter optimizer"""
        messagebox.showinfo("Coming Soon", "Parameter optimization tool")

    def run_script_mode(self):
        """Run in script mode"""
        messagebox.showinfo("Coming Soon", "Launch script mode analysis")

    def show_quick_start(self):
        """Show quick start guide"""
        messagebox.showinfo("Quick Start",
            "1. Select a preset from Configuration tab\n"
            "2. Adjust parameters as needed\n"
            "3. Click 'Run Analysis'\n"
            "4. View results in Analysis Results tab\n"
            "5. Create custom plots in Interactive Plots tab")

    def show_documentation(self):
        """Show documentation"""
        messagebox.showinfo("Documentation", "Opening README_v3.md...")

    def show_parameter_guide(self):
        """Show parameter tuning guide"""
        messagebox.showinfo("Parameter Guide", "See example_v3_mission_analysis.py Example 4")

    def show_about(self):
        """Show about dialog"""
        messagebox.showinfo("About",
            "VTOL Performance Analyzer v4.0\n"
            "Professional Edition\n\n"
            "Industrial-grade tailsitter performance analysis\n"
            "with full GUI and interactive tools.\n\n"
            "Version: 4.0.0\n"
            "Date: 2025-01-20")

    # -----------------------------------------------------------------------
    # PRESET ACTIONS
    # -----------------------------------------------------------------------

    def on_preset_selected(self, event=None):
        """Handle preset selection from dropdown"""
        pass  # User must click Load button

    def load_selected_preset(self):
        """Load the selected preset"""
        selected = self.preset_var.get()
        preset_name = selected.split(':')[0]
        self.load_preset(preset_name)

    def save_preset_as(self):
        """Save current configuration as new preset"""
        messagebox.showinfo("Coming Soon", "Save as custom preset")

    def reset_config(self):
        """Reset to current preset defaults"""
        if messagebox.askyesno("Reset", "Reset to preset defaults?"):
            self.load_preset(self.current_preset_name)

    def apply_config(self):
        """Apply parameter changes"""
        try:
            # Update config from UI widgets
            self.update_config_from_ui()

            # Mark as modified
            self.mark_config_modified()

            self.update_status("Configuration updated")

            if self.auto_update.get():
                self.run_analysis()
        except Exception as e:
            messagebox.showerror("Error", f"Could not apply configuration: {e}")

    def update_config_from_ui(self):
        """Update configuration from UI parameter widgets"""
        if not self.current_config:
            return

        try:
            # Basic parameters
            if 'total_takeoff_weight_kg' in self.param_widgets:
                self.current_config.total_takeoff_weight_kg = float(self.param_widgets['total_takeoff_weight_kg'].get())
            if 'wingspan_m' in self.param_widgets:
                self.current_config.wingspan_m = float(self.param_widgets['wingspan_m'].get())
            if 'wing_chord_m' in self.param_widgets:
                self.current_config.wing_chord_m = float(self.param_widgets['wing_chord_m'].get())
            if 'field_elevation_m' in self.param_widgets:
                self.current_config.field_elevation_m = float(self.param_widgets['field_elevation_m'].get())

            # Tailsitter parameters
            if 'control_power_base_w' in self.param_widgets:
                self.current_config.control_power_base_w = float(self.param_widgets['control_power_base_w'].get())
            if 'control_power_speed_factor' in self.param_widgets:
                self.current_config.control_power_speed_factor = float(self.param_widgets['control_power_speed_factor'].get())

            # Drag coefficients
            if 'cd0_motor_nacelles' in self.param_widgets:
                self.current_config.cd0_motor_nacelles = float(self.param_widgets['cd0_motor_nacelles'].get())
            if 'cd0_fuselage_base' in self.param_widgets:
                self.current_config.cd0_fuselage_base = float(self.param_widgets['cd0_fuselage_base'].get())
            if 'cd0_landing_gear' in self.param_widgets:
                self.current_config.cd0_landing_gear = float(self.param_widgets['cd0_landing_gear'].get())
            if 'cd0_interference' in self.param_widgets:
                self.current_config.cd0_interference = float(self.param_widgets['cd0_interference'].get())

            # Transitions
            if 'transition_forward_duration_s' in self.param_widgets:
                self.current_config.transition_forward_duration_s = float(self.param_widgets['transition_forward_duration_s'].get())
            if 'transition_forward_power_factor' in self.param_widgets:
                self.current_config.transition_forward_power_factor = float(self.param_widgets['transition_forward_power_factor'].get())
            if 'transition_back_duration_s' in self.param_widgets:
                self.current_config.transition_back_duration_s = float(self.param_widgets['transition_back_duration_s'].get())
            if 'transition_back_power_factor' in self.param_widgets:
                self.current_config.transition_back_power_factor = float(self.param_widgets['transition_back_power_factor'].get())

            # Propulsion efficiencies (convert from percentage)
            if 'prop_efficiency_lowspeed' in self.param_widgets:
                self.current_config.prop_efficiency_lowspeed = float(self.param_widgets['prop_efficiency_lowspeed'].get()) / 100.0
            if 'prop_efficiency_highspeed' in self.param_widgets:
                self.current_config.prop_efficiency_highspeed = float(self.param_widgets['prop_efficiency_highspeed'].get()) / 100.0
            if 'motor_efficiency_peak' in self.param_widgets:
                self.current_config.motor_efficiency_peak = float(self.param_widgets['motor_efficiency_peak'].get()) / 100.0
            if 'esc_efficiency' in self.param_widgets:
                self.current_config.esc_efficiency = float(self.param_widgets['esc_efficiency'].get()) / 100.0

            # Auxiliary systems
            if 'avionics_power_w' in self.param_widgets:
                self.current_config.avionics_power_w = float(self.param_widgets['avionics_power_w'].get())
            if 'payload_power_w' in self.param_widgets:
                self.current_config.payload_power_w = float(self.param_widgets['payload_power_w'].get())
            if 'heater_power_w' in self.param_widgets:
                self.current_config.heater_power_w = float(self.param_widgets['heater_power_w'].get())

            # Geometry parameters (v4.1)
            if 'fuselage_length_m' in self.param_widgets:
                self.current_config.fuselage_length_m = float(self.param_widgets['fuselage_length_m'].get())
            if 'fuselage_diameter_m' in self.param_widgets:
                self.current_config.fuselage_diameter_m = float(self.param_widgets['fuselage_diameter_m'].get())
            if 'num_tail_fins' in self.param_widgets:
                self.current_config.num_tail_fins = int(self.param_widgets['num_tail_fins'].get())
            if 'tail_fin_chord_m' in self.param_widgets:
                self.current_config.tail_fin_chord_m = float(self.param_widgets['tail_fin_chord_m'].get())
            if 'tail_fin_span_m' in self.param_widgets:
                self.current_config.tail_fin_span_m = float(self.param_widgets['tail_fin_span_m'].get())
            if 'tail_fin_position_m' in self.param_widgets:
                self.current_config.tail_fin_position_m = float(self.param_widgets['tail_fin_position_m'].get())
            if 'tail_fin_thickness_ratio' in self.param_widgets:
                self.current_config.tail_fin_thickness_ratio = float(self.param_widgets['tail_fin_thickness_ratio'].get())
            if 'tail_fin_taper_ratio' in self.param_widgets:
                self.current_config.tail_fin_taper_ratio = float(self.param_widgets['tail_fin_taper_ratio'].get())
            if 'motor_spacing_m' in self.param_widgets:
                self.current_config.motor_spacing_m = float(self.param_widgets['motor_spacing_m'].get())
            if 'num_motors' in self.param_widgets:
                self.current_config.num_motors = int(self.param_widgets['num_motors'].get())

            # Recalculate derived parameters
            self.current_config.__post_init__()

        except ValueError as e:
            raise ValueError(f"Invalid parameter value: {e}")

    def get_current_config(self):
        """Get current configuration from UI (v4.1)"""
        # Update config from UI values
        self.update_config_from_ui()
        # Return the updated config
        return self.current_config

    def validate_config(self):
        """Validate current configuration"""
        try:
            errors = []
            warnings = []

            # Validate basic parameters
            if self.current_config.total_takeoff_weight_kg < 1.0 or self.current_config.total_takeoff_weight_kg > 20.0:
                errors.append("Total weight must be between 1.0 and 20.0 kg")

            if self.current_config.wingspan_m < 0.5 or self.current_config.wingspan_m > 5.0:
                errors.append("Wing span must be between 0.5 and 5.0 m")

            if self.current_config.wing_chord_m < 0.05 or self.current_config.wing_chord_m > 0.5:
                errors.append("Wing chord must be between 0.05 and 0.5 m")

            # Validate efficiencies
            if self.current_config.prop_efficiency_lowspeed < 0.3 or self.current_config.prop_efficiency_lowspeed > 1.0:
                errors.append("Propeller efficiency (low speed) must be between 30% and 100%")

            if self.current_config.motor_efficiency_peak < 0.5 or self.current_config.motor_efficiency_peak > 1.0:
                errors.append("Motor efficiency must be between 50% and 100%")

            # Validate drag coefficients
            if self.current_config.cd0_motor_nacelles < 0.01 or self.current_config.cd0_motor_nacelles > 0.1:
                warnings.append("CD0 motor nacelles seems unusual (typical: 0.020-0.050)")

            # Validate power parameters
            if self.current_config.control_power_base_w < 10 or self.current_config.control_power_base_w > 200:
                warnings.append("Control power base seems unusual (typical: 30-100 W)")

            # Display results
            if errors:
                messagebox.showerror("Validation Failed",
                    "Configuration has errors:\n\n" + "\n".join(f"• {e}" for e in errors))
                return False
            elif warnings:
                response = messagebox.showwarning("Validation Warnings",
                    "Configuration has warnings:\n\n" + "\n".join(f"• {w}" for w in warnings) +
                    "\n\nDo you want to continue?",
                    type=messagebox.OKCANCEL)
                return response == 'ok'
            else:
                messagebox.showinfo("Validation Passed",
                    "✓ Configuration is valid!\n\n"
                    "All parameters are within acceptable ranges.")
                return True

        except Exception as e:
            messagebox.showerror("Validation Error", f"Could not validate configuration:\n{e}")
            return False

    # -----------------------------------------------------------------------
    # PLOTTING FUNCTIONS
    # -----------------------------------------------------------------------

    def get_plottable_parameters(self):
        """Get list of parameters that can be plotted"""
        return [
            "Speed (m/s)",
            "Forward Flight Power (W)",
            "Hover Power (W)",
            "Control Power (W)",
            "Current (A)",
            "Forward Flight Endurance (min)",
            "Hover Endurance (pure hover) (min)",
            "Forward Flight Range (km)",
            "Max Range (optimized speed) (km)",
            "Weight (kg)",
            "Wing Span (m)",
            "Wing Area (m²)",
            "Altitude (m)",
            "Battery Capacity (mAh)",
            "Propeller Efficiency (%)",
            "Lift-to-Drag Ratio",
        ]

    def add_plot_parameter(self, param_name=None):
        """Add a parameter to the plot selection list"""
        # Limit to 3 parameters
        if len(self.plot_params) >= 3:
            messagebox.showwarning("Maximum Reached", "You can select up to 3 parameters")
            return

        # Create parameter row
        param_row = ttk.Frame(self.params_container)
        param_row.pack(fill='x', pady=3)

        # Parameter number label
        param_num = len(self.plot_params) + 1
        ttk.Label(param_row, text=f"Parameter {param_num}:", width=12).pack(side='left', padx=5)

        # Parameter dropdown
        param_var = tk.StringVar(value=param_name if param_name else "Speed (m/s)")
        param_combo = ttk.Combobox(
            param_row,
            textvariable=param_var,
            values=self.get_plottable_parameters(),
            width=35,
            state='readonly'
        )
        param_combo.pack(side='left', padx=5)

        # Remove button
        remove_btn = ttk.Button(
            param_row,
            text="➖",
            width=3,
            command=lambda: self.remove_plot_parameter(param_row, param_data)
        )
        remove_btn.pack(side='left', padx=5)

        # Store parameter data
        param_data = {
            'frame': param_row,
            'var': param_var,
            'combo': param_combo,
        }
        self.plot_params.append(param_data)

    def remove_plot_parameter(self, frame, param_data):
        """Remove a parameter from the plot selection list"""
        # Remove from UI
        frame.destroy()

        # Remove from list
        if param_data in self.plot_params:
            self.plot_params.remove(param_data)

        # Renumber remaining parameters
        for i, param in enumerate(self.plot_params):
            # Update label
            for widget in param['frame'].winfo_children():
                if isinstance(widget, ttk.Label):
                    widget.config(text=f"Parameter {i+1}:")
                    break

    def clear_all_plot_params(self):
        """Clear all plot parameters"""
        for param in self.plot_params[:]:  # Copy list to avoid modification during iteration
            param['frame'].destroy()
        self.plot_params = []
        self.clear_plot()

    def load_common_plot(self, plot_id):
        """Load a common plot configuration"""
        try:
            from plots import COMMON_PLOTS

            plot_def = COMMON_PLOTS.get(plot_id)
            if not plot_def:
                messagebox.showerror("Error", f"Plot definition not found: {plot_id}")
                return

            # Clear existing parameters
            self.clear_all_plot_params()

            # Add X parameter
            self.add_plot_parameter(plot_def['x_param'])

            # Add Y parameters
            for y_param in plot_def['y_params']:
                self.add_plot_parameter(y_param)

            # Generate plot automatically
            self.after(100, self.generate_custom_plot)

            self.update_status(f"✓ Loaded: {plot_def['name']}")

        except Exception as e:
            messagebox.showerror("Load Error", f"Could not load common plot:\n{e}")

    def create_tooltip(self, widget, text):
        """Create a tooltip for a widget"""
        def on_enter(event):
            tooltip = tk.Toplevel()
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")

            label = ttk.Label(
                tooltip,
                text=text,
                background="#ffffe0",
                relief='solid',
                borderwidth=1,
                font=('Arial', 9)
            )
            label.pack()

            widget.tooltip = tooltip

        def on_leave(event):
            if hasattr(widget, 'tooltip'):
                widget.tooltip.destroy()
                del widget.tooltip

        widget.bind("<Enter>", on_enter)
        widget.bind("<Leave>", on_leave)

    def generate_custom_plot(self):
        """Generate user-defined custom plot with dynamic parameters"""
        if not self.current_calc:
            messagebox.showwarning("No Analysis", "Run an analysis first!")
            return

        # Validate parameter count
        if len(self.plot_params) < 2:
            messagebox.showwarning("Not Enough Parameters",
                                 "Add at least 2 parameters to create a plot.\n\n"
                                 "2 params: Y vs X\n"
                                 "3 params: X vs Y1, Y2")
            return

        try:
            self.update_status("Generating plot...")

            # Clear existing plot
            self.clear_plot()

            # Get parameter selections
            params = [p['var'].get() for p in self.plot_params]

            if len(params) == 2:
                # 2 parameters: param2 vs param1 (Y vs X)
                x_param = params[0]
                y_params = [params[1]]
                title = f"{params[1]} vs {params[0]}"

            elif len(params) == 3:
                # 3 parameters: param1 vs param2, param3 (X vs Y1, Y2)
                x_param = params[0]
                y_params = [params[1], params[2]]
                title = f"{params[0]} vs {params[1]}, {params[2]}"

            # Create matplotlib figure
            fig = Figure(figsize=(10, 6), dpi=100)
            ax = fig.add_subplot(111)

            # Colors for multiple lines
            colors = ['#2E86AB', '#A23B72', '#F18F01']

            # Plot each Y parameter
            all_y_data = []
            for i, y_param in enumerate(y_params):
                x_data, y_data, x_label, y_label = self.calculate_plot_data(x_param, y_param)

                ax.plot(x_data, y_data, color=colors[i], linewidth=2,
                       label=y_label, marker='o', markersize=3, markevery=5)

                all_y_data.append((y_data, y_label))

            # Set labels and formatting
            ax.set_xlabel(x_label, fontsize=11, fontweight='bold')

            if len(y_params) == 1:
                ax.set_ylabel(y_label, fontsize=11, fontweight='bold')
            else:
                ax.set_ylabel("Value", fontsize=11, fontweight='bold')

            ax.set_title(title, fontsize=13, fontweight='bold', pad=15)
            ax.grid(True, alpha=0.3, linestyle='--')
            ax.legend(loc='best', framealpha=0.9)

            # Tight layout
            fig.tight_layout()

            # Embed in tkinter
            canvas = FigureCanvasTkAgg(fig, master=self.plot_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True)

            # Add toolbar
            toolbar = NavigationToolbar2Tk(canvas, self.plot_frame)
            toolbar.update()

            # Store for export
            self.current_plot_fig = fig
            self.current_plot_data = (x_data, all_y_data, x_label)

            self.update_status("✓ Plot generated successfully")

        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Plot Error", f"Could not generate plot:\n{e}")
            self.update_status("Plot generation failed")

    def calculate_plot_data(self, x_param, y_param):
        """Calculate data for plotting based on parameter selection"""
        import numpy as np

        # Determine sweep range based on x parameter
        if "Speed" in x_param:
            x_range = np.linspace(10, 25, 50)
            x_label = "Speed (m/s)"
        elif "Weight" in x_param:
            x_range = np.linspace(4, 8, 50)
            x_label = "Weight (kg)"
        elif "Wing Span" in x_param:
            x_range = np.linspace(1.5, 2.5, 50)
            x_label = "Wing Span (m)"
        elif "Wing Area" in x_param:
            x_range = np.linspace(0.2, 0.6, 50)
            x_label = "Wing Area (m²)"
        elif "Altitude" in x_param:
            x_range = np.linspace(0, 3000, 50)
            x_label = "Altitude (m MSL)"
        elif "Battery Capacity" in x_param:
            x_range = np.linspace(5000, 20000, 50)
            x_label = "Battery Capacity (mAh)"
        elif "Control Power" in x_param:
            x_range = np.linspace(20, 100, 50)
            x_label = "Control Power (W)"
        else:
            # Default sweep
            x_range = np.linspace(10, 25, 50)
            x_label = x_param

        x_values = x_range

        # Calculate y values for each x
        y_values = []
        for x_val in x_range:
            y_val = self.calculate_y_for_x(x_param, x_val, y_param)
            y_values.append(y_val)

        y_values = np.array(y_values)
        y_label = y_param

        return x_values, y_values, x_label, y_label

    def calculate_y_for_x(self, x_param, x_val, y_param):
        """Calculate Y value for given X parameter value"""
        from dataclasses import replace

        temp_config = replace(self.current_config)

        # Modify config based on x parameter
        if "Speed" in x_param:
            speed = x_val
        elif "Weight" in x_param:
            temp_config.total_takeoff_weight_kg = x_val
            temp_config.__post_init__()
            speed = 15.0
        elif "Wing Span" in x_param:
            temp_config.wingspan_m = x_val
            temp_config.__post_init__()
            speed = 15.0
        elif "Wing Area" in x_param:
            temp_config.wing_area_m2 = x_val
            temp_config.__post_init__()
            speed = 15.0
        elif "Altitude" in x_param:
            temp_config.field_elevation_m = x_val
            temp_config.__post_init__()
            speed = 15.0
        elif "Battery Capacity" in x_param:
            temp_config.battery_capacity_mah = x_val
            temp_config.__post_init__()
            speed = 15.0
        elif "Control Power" in x_param:
            temp_config.control_power_w = x_val
            temp_config.__post_init__()
            speed = 15.0
        else:
            speed = 15.0

        # Calculate based on y parameter
        temp_calc = PerformanceCalculator(temp_config)

        if "Forward Flight Power" in y_param:
            pb = temp_calc.power_budget_breakdown(speed)
            return pb['total_electrical_w']
        elif "Hover Power" in y_param:
            return temp_calc.hover_power()
        elif "Control Power" in y_param:
            return temp_calc.control_power(speed)
        elif "Current" in y_param:
            return temp_calc.cruise_current(speed)
        elif "Forward Flight Endurance" in y_param:
            current = temp_calc.cruise_current(speed)
            return temp_calc.endurance(current)
        elif "Hover Endurance" in y_param:
            current = temp_calc.hover_current()
            return temp_calc.endurance(current)
        elif "Forward Flight Range" in y_param and "Max Range" not in y_param:
            current = temp_calc.cruise_current(speed)
            endurance = temp_calc.endurance(current)
            return temp_calc.range_km(speed, endurance)
        elif "Max Range" in y_param:
            best_range_result = temp_calc.best_range_speed()
            return best_range_result['range_km']
        elif "Propeller Efficiency" in y_param:
            return temp_calc.propeller_efficiency_cruise(speed) * 100
        elif "Lift-to-Drag Ratio" in y_param:
            return temp_calc.lift_to_drag_ratio(speed)
        else:
            return 0.0

    def clear_plot(self):
        """Clear current plot"""
        for widget in self.plot_frame.winfo_children():
            widget.destroy()

        self.current_plot_fig = None
        self.current_plot_data = None

    def export_plot_png(self):
        """Export current plot as PNG"""
        if not hasattr(self, 'current_plot_fig') or self.current_plot_fig is None:
            messagebox.showwarning("No Plot", "Generate a plot first!")
            return

        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".png",
                filetypes=[("PNG files", "*.png"), ("All files", "*.*")]
            )

            if filename:
                self.current_plot_fig.savefig(filename, dpi=300, bbox_inches='tight')
                self.update_status(f"Plot saved to {filename}")
                messagebox.showinfo("Success", f"Plot saved to:\n{filename}")

        except Exception as e:
            messagebox.showerror("Export Error", f"Could not export plot:\n{e}")

    def export_plot_csv(self):
        """Export plot data as CSV"""
        if not hasattr(self, 'current_plot_data') or self.current_plot_data is None:
            messagebox.showwarning("No Data", "Generate a plot first!")
            return

        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
            )

            if filename:
                import csv
                x_data, all_y_data, x_label = self.current_plot_data

                with open(filename, 'w', newline='') as f:
                    writer = csv.writer(f)

                    # Header row
                    headers = [x_label] + [y_label for _, y_label in all_y_data]
                    writer.writerow(headers)

                    # Data rows
                    for i, x in enumerate(x_data):
                        row = [x] + [y_data[i] for y_data, _ in all_y_data]
                        writer.writerow(row)

                self.update_status(f"✓ Data saved to {filename}")
                messagebox.showinfo("Success", f"Data saved to:\n{filename}")

        except Exception as e:
            messagebox.showerror("Export Error", f"Could not export data:\n{e}")

    def quick_plot(self, x_param, y_param):
        """Generate a quick plot"""
        # Set dropdowns
        self.plot_x_var.set(f"{x_param} (m/s)" if x_param == "Speed" else f"{x_param} (kg)")
        self.plot_y_var.set(f"{y_param} (W)" if y_param == "Power" else f"{y_param} (km)" if y_param == "Range" else f"{y_param} (min)")

        # Generate
        self.generate_custom_plot()

    # -----------------------------------------------------------------------
    # EXPORT FUNCTIONS
    # -----------------------------------------------------------------------

    def export_results_pdf(self):
        """Export results as PDF"""
        messagebox.showinfo("Coming Soon", "Export results as PDF")

    def copy_results_text(self):
        """Copy results text to clipboard"""
        if self.current_results:
            self.clipboard_clear()
            self.clipboard_append(self.results_text.get('1.0', tk.END))
            self.update_status("Results copied to clipboard")
        else:
            messagebox.showinfo("No Results", "Run an analysis first")

    # -----------------------------------------------------------------------
    # CONFIGURATION PERSISTENCE
    # -----------------------------------------------------------------------

    def save_config_to_file(self, filename):
        """Save current configuration to JSON file"""
        import json
        from dataclasses import asdict

        config_dict = asdict(self.current_config)

        # Add metadata
        data = {
            "metadata": {
                "preset_name": self.current_preset_name,
                "version": "4.0",
                "timestamp": self.get_timestamp()
            },
            "configuration": config_dict
        }

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2)

    def load_config_from_file(self, filename):
        """Load configuration from JSON file"""
        import json

        with open(filename, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # Extract configuration
        config_dict = data.get('configuration', data)  # Backward compatible

        # Update current config
        for key, value in config_dict.items():
            if hasattr(self.current_config, key):
                setattr(self.current_config, key, value)

        # Recalculate derived parameters
        self.current_config.__post_init__()

        # Update UI
        self.update_ui_from_config()

        # Update preset name if available
        metadata = data.get('metadata', {})
        if 'preset_name' in metadata:
            self.current_preset_name = metadata['preset_name']
            self.preset_status.set(f"Preset: {self.current_preset_name.upper()} (loaded)")

    def add_to_recent(self, filename):
        """Add file to recent configurations list"""
        # Remove if already in list
        if filename in self.recent_configs:
            self.recent_configs.remove(filename)

        # Add to beginning
        self.recent_configs.insert(0, filename)

        # Trim to max
        self.recent_configs = self.recent_configs[:self.max_recent]

        # Update menu
        self.update_recent_menu()

    def update_recent_menu(self):
        """Update recent files menu (if it exists)"""
        # This would update a "Recent Files" submenu if we had one
        pass

    def get_timestamp(self):
        """Get current timestamp"""
        from datetime import datetime
        return datetime.now().isoformat()

    # -----------------------------------------------------------------------
    # AUTO-SAVE FUNCTIONALITY
    # -----------------------------------------------------------------------

    def start_auto_save_timer(self):
        """Start auto-save timer (saves every 5 minutes)"""
        if self.auto_save_timer_id:
            self.after_cancel(self.auto_save_timer_id)

        # Auto-save every 5 minutes (300000 ms)
        self.auto_save_timer_id = self.after(300000, self.auto_save)

    def auto_save(self):
        """Perform auto-save"""
        if self.config_modified and self.auto_save_enabled.get():
            try:
                auto_save_file = os.path.join(self.config_dir, "autosave.json")
                self.save_config_to_file(auto_save_file)
                self.update_status("Auto-saved configuration")
            except:
                pass  # Silently fail auto-save

        # Restart timer
        if self.auto_save_enabled.get():
            self.start_auto_save_timer()

    def mark_config_modified(self):
        """Mark configuration as modified"""
        if not self.config_modified:
            self.config_modified = True
            # Update window title to show modification
            current_title = self.title()
            if not current_title.endswith("*"):
                self.title(current_title + " *")

    # -----------------------------------------------------------------------
    # SESSION STATE MANAGEMENT
    # -----------------------------------------------------------------------

    def save_session_state(self):
        """Save current session state"""
        import json

        try:
            session_data = {
                "version": "4.0",
                "timestamp": self.get_timestamp(),
                "last_preset": self.current_preset_name,
                "window_geometry": self.geometry(),
                "auto_save_enabled": self.auto_save_enabled.get(),
                "auto_update": self.auto_update.get(),
                "recent_configs": self.recent_configs,
                "output_directory": self.output_dir_var.get() if hasattr(self, 'output_dir_var') else "./output"
            }

            session_file_path = os.path.join(self.config_dir, self.session_file)
            with open(session_file_path, 'w', encoding='utf-8') as f:
                json.dump(session_data, f, indent=2)

        except Exception as e:
            print(f"Could not save session state: {e}")

    def load_session_state(self):
        """Load previous session state"""
        import json

        try:
            session_file_path = os.path.join(self.config_dir, self.session_file)

            if os.path.exists(session_file_path):
                with open(session_file_path, 'r', encoding='utf-8') as f:
                    session_data = json.load(f)

                # Restore settings
                if 'last_preset' in session_data:
                    self.load_preset(session_data['last_preset'])

                if 'auto_save_enabled' in session_data:
                    self.auto_save_enabled.set(session_data['auto_save_enabled'])

                if 'auto_update' in session_data:
                    self.auto_update.set(session_data['auto_update'])

                if 'recent_configs' in session_data:
                    self.recent_configs = session_data['recent_configs']
                    # Verify files still exist
                    self.recent_configs = [f for f in self.recent_configs if os.path.exists(f)]

                if 'output_directory' in session_data and hasattr(self, 'output_dir_var'):
                    self.output_dir_var.set(session_data['output_directory'])

                # Try to restore window geometry
                # if 'window_geometry' in session_data:
                #     self.geometry(session_data['window_geometry'])

                self.update_status("Session restored")

            else:
                # First run - load default
                self.load_preset("baseline")

        except Exception as e:
            print(f"Could not load session state: {e}")
            # Fallback to default
            self.load_preset("baseline")

    def check_for_autosave(self):
        """Check if there's an autosave file and ask to restore"""
        auto_save_file = os.path.join(self.config_dir, "autosave.json")

        if os.path.exists(auto_save_file):
            # Check if it's newer than 1 hour
            import time
            file_time = os.path.getmtime(auto_save_file)
            current_time = time.time()

            if (current_time - file_time) < 3600:  # Within last hour
                response = messagebox.askyesno(
                    "Auto-Save Found",
                    "An auto-saved configuration was found.\nWould you like to restore it?"
                )

                if response:
                    try:
                        self.load_config_from_file(auto_save_file)
                        self.update_status("Restored from auto-save")
                    except:
                        pass

    # -----------------------------------------------------------------------
    # WINDOW CLOSE
    # -----------------------------------------------------------------------

    def on_closing(self):
        """Handle window close"""
        # Check for unsaved changes
        if self.config_modified:
            response = messagebox.askyesnocancel(
                "Unsaved Changes",
                "Save configuration before closing?"
            )

            if response is None:  # Cancel
                return
            elif response:  # Yes, save
                self.save_config()

        # Save session state
        self.save_session_state()

        # Cancel auto-save timer
        if self.auto_save_timer_id:
            self.after_cancel(self.auto_save_timer_id)

        # Close application
        self.destroy()


# ===========================================================================
# MAIN ENTRY POINT
# ===========================================================================

def main():
    """Launch GUI application"""
    app = VTOLAnalyzerGUI()
    app.mainloop()


if __name__ == "__main__":
    main()
